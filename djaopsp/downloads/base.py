# Copyright (c) 2024, DjaoDjin inc.
# see LICENSE.
from __future__ import unicode_literals

import csv, io, math

# importing `survey.mixins.TimersMixin` results in an import loop:
#   cannot import name 'get_object_or_404' from 'rest_framework.generics'??
from deployutils.apps.django_deployutils.mixins.timers import TimersMixin
from rest_framework.renderers import BaseRenderer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter
from pages.helpers import get_extra

from ..compat import force_str, six
from ..helpers import as_valid_sheet_title


class CSVDownloadRenderer(BaseRenderer):
    """
    As CVS file
    """
    media_type = 'text/csv'
    format = 'csv'

    @staticmethod
    def encode(text):
        if six.PY2:
            return force_str(text).encode('utf-8')
        return force_str(text)

    def get_headings(self, renderer_context=None):
        if renderer_context:
            view = renderer_context.get('view')
            return view.headings
        return []

    def format_row(self, entry):
        row = []
        # XXX Surprisingly the values are iterated in the correct order.
        for field_value in six.itervalues(entry):
            row += [field_value]
        return row

    def render(self, data, accepted_media_type=None, renderer_context=None):
        if six.PY2:
            content = io.BytesIO()
        else:
            content = io.StringIO()
        csv_writer = csv.writer(content)
        headings = self.get_headings(renderer_context=renderer_context)
        csv_writer.writerow([self.encode(head) for head in headings])
        if isinstance(data, dict):
            results = data.get('results', [])
        else:
            results = data
        for entry in results:
            csv_writer.writerow(self.format_row(entry))
        content.seek(0)
        return content


class XLSXRenderer(TimersMixin, BaseRenderer):

    media_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'

    def flush_writer(self, wbook):
        """
        Write out the Excel file.
        """
        content = io.BytesIO()
        wbook.save(content)
        content.seek(0)
        return content

    def format_row(self, entry):
        row = []
        for field_value in six.itervalues(entry):
            row += [str(field_value)]
        return row

    def get_headings(self, renderer_context=None):
        if renderer_context:
            view = renderer_context.get('view')
            return view.headings
        return []

    @staticmethod
    def get_indent_question(depth=0):
        return "  " * depth

    @staticmethod
    def get_indent_heading(depth=0):
        return "  " * depth

    def render(self, data, accepted_media_type=None, renderer_context=None):
        wbook = Workbook()

        title = None
        descr = None
        if renderer_context:
            view = renderer_context.get('view')
            title = view.title
            descr = view.descr

        # Populate the Total sheet
        wsheet = wbook.active

        if title:
            wsheet.title = as_valid_sheet_title(title)
        if descr:
            wsheet.append([descr])
        wsheet.append(self.get_headings(renderer_context=renderer_context))

        if isinstance(data, dict):
            results = data.get('results', [])
        else:
            results = data
        for entry in results:
            wsheet.append(self.format_row(entry))

        # Prepares the result file
        content = io.BytesIO()
        wbook.save(content)
        content.seek(0)

        return self.flush_writer(wbook)


class PracticesXLSXRenderer(TimersMixin, BaseRenderer):

    add_style = True
    add_expanded_styles = True
    depth = 0
    basename = 'practices'
    media_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    base_headers = ['']
    peer_value_headers = []
    intrinsic_value_headers = []

    valueFills = [
        PatternFill(fill_type=FILL_SOLID, fgColor='FF9CD76B'), # green-level-0
        PatternFill(fill_type=FILL_SOLID, fgColor='FF69B02B'), # green-level-1
        PatternFill(fill_type=FILL_SOLID, fgColor='FF007C3F'), # green-level-2
        PatternFill(fill_type=FILL_SOLID, fgColor='FFFFD700'), # green-level-3
    ]

    def __init__(self, *args):
        super(PracticesXLSXRenderer, self).__init__(*args)
        self.wbook = None
        self.wsheet = None

    def get_title_row(self):
        return []

    def get_value_fill(self, val):
        idx = val - 1
        if idx < len(self.valueFills):
            return self.valueFills[idx]
        return self.valueFills[-1]

    def get_headings(self, renderer_context=None):
        view = renderer_context.get('view')
        base_headers = view.base_headers
        intrinsic_value_headers = view.intrinsic_value_headers
        headers = [] + base_headers
        headers += intrinsic_value_headers
        return headers


    def format_row(self, entry, key=None):
        #pylint:disable=unused-argument
        # base_headers
        title = entry['title']
        row = [title]

        # intrinsic_value_headers
        avg_value = 0
        extra = entry.get('extra')
        if extra:
            intrinsic_values = extra.get('intrinsic_values')
            if intrinsic_values:
                environmental_value = intrinsic_values.get('environmental', 0)
                business_value = intrinsic_values.get('business', 0)
                profitability = intrinsic_values.get('profitability', 0)
                implementation_ease = intrinsic_values.get(
                    'implementation_ease', 0)
                avg_value = (environmental_value + business_value +
                    profitability + implementation_ease) // 4
        if avg_value:
            row += [
                environmental_value,
                business_value,
                profitability,
                implementation_ease,
                avg_value
            ]
        else:
            row += ['', '', '', '', '']
        return row

    def is_practice(self, entry):
        default_unit = entry.get('default_unit')
        intrinsic_values = get_extra(entry, 'intrinsic_values')
        return default_unit or intrinsic_values

    def create_writer(self, title=None):
        #pylint:disable=unused-argument
        if not hasattr(self, 'wbook') or not self.wbook:
            self.wbook = Workbook()
            self.wsheet = self.wbook.active
            if title:
                self.wsheet.title = as_valid_sheet_title(title)
        else:
            self.wsheet = self.wbook.create_sheet(
                as_valid_sheet_title(title))

    def flush_writer(self, wbook):
        self.optimal_cell_sizes()
        # Write out the Excel file.
        content = io.BytesIO()
        wbook.save(content)
        content.seek(0)
        return content

    def write_headers(self, renderer_context=None):
        """
        Write table headers in the worksheet
        """
        view = renderer_context.get('view')
        intrinsic_value_headers = view.intrinsic_value_headers

        self.wsheet.append(self.get_title_row())
        super_headers = []
        nb_peer_value_headers = 0
        nb_intrinsic_value_headers = 0
        # We have to call `get_headings`
        # before retrieving `peer_value_headers`.
        headers = self.get_headings(renderer_context=renderer_context)
        if self.peer_value_headers:
            nb_peer_value_headers = 1
            super_headers += ['Peer-based value'] # for correct nb cols
            for header in self.peer_value_headers:
                super_headers += ["" for unused in range(0, len(header[1]))]
                nb_peer_value_headers += len(header[1])
        if intrinsic_value_headers:
            nb_intrinsic_value_headers = len(intrinsic_value_headers)
            super_headers += ['Intrinsic value'] + [
                "" for unused in range(1, nb_intrinsic_value_headers)]
        if super_headers:
            super_headers = ([
                "" for unused in range(0, len(self.base_headers))] +
                super_headers)
            self.wsheet.append(super_headers)
            if nb_peer_value_headers:
                first_col = chr(ord('A') + len(self.base_headers))
                last_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers - 1)
                self.wsheet.merge_cells('%s1:%s1' % (first_col, last_col))
            if nb_intrinsic_value_headers:
                first_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers)
                last_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers + nb_intrinsic_value_headers - 1)
                self.wsheet.merge_cells('%s1:%s1' % (first_col, last_col))
        self.wsheet.append(headers)

    def write_sheet(self, queryset=None, title="", key=None,
                    renderer_context=None):
        #pylint:disable=too-many-locals,too-many-nested-blocks
        view = renderer_context.get('view')
        intrinsic_value_headers = view.intrinsic_value_headers

        self.create_writer(title=title)
        self.write_headers(renderer_context=renderer_context)
        self._report_queries(
            "headers written in sheet '%s'" % title)
        start_row = self.wsheet.max_row
        for entry in queryset:
            self.wsheet.append(self.format_row(entry, key=key))
        self._report_queries("rows written in sheet '%s'" % title)

        if self.add_style:
            max_indent = 0
            height_ratio = 2.8
            mm_to_pts_ratio = 0.5102
            max_heading_width = 132.15 * mm_to_pts_ratio
            row_heights = []
            for idx, entry in enumerate(queryset):
                max_indent = max(max_indent, entry.get('indent', 0))
                #max_heading_width = max(max_heading_width,
                #    entry['indent'] + len(entry['title']))
                nb_wrapped_lines = max(int((
                    entry['indent'] + len(entry['title']))
                    * 0.9 / max_heading_width) + 1, 1)
                row_heights += [nb_wrapped_lines * 5.29 * height_ratio]

            title_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            heading_font = Font(
                name='Calibri', size=12, bold=False, italic=False,
                vertAlign='baseline', underline='none', strike=False,
                color='54BAD8')
            practice_font = Font(
                name='Calibri', size=12, bold=False, italic=False,
                vertAlign='baseline', underline='none', strike=False,
                color='777777')
            tile_background = PatternFill(
                fill_type=FILL_SOLID, fgColor='FFFFA6')
            bordered = Border(
                left=Side(border_style=BORDER_THIN, color='FF000000'),
                right=Side(border_style=BORDER_THIN, color='FF000000'),
                top=Side(border_style=BORDER_THIN, color='FF000000'),
                bottom=Side(border_style=BORDER_THIN, color='FF000000'))
            inner_cell_alignment = Alignment(
                horizontal="left", vertical='top', wrap_text=True)
            # Creates a table of indentation for row titles
            first_col_alignments = [Alignment(horizontal='left', vertical='top',
                            indent=indent, wrap_text=True)
                for indent in range(0, max_indent + 1)]

            self._report_queries("computed styles for sheet '%s'" % title)

            # Column headers
            for column in self.wsheet.columns:
                try:
                    column_letter = column[0].column_letter
                    self.wsheet.column_dimensions[column_letter].width = \
                        22.58 * mm_to_pts_ratio
                except AttributeError:
                    continue
            self.wsheet.column_dimensions['A'].width = max_heading_width
            for row_cells in self.wsheet.iter_rows(
                    min_row=0, max_row=start_row):
                for cell in row_cells:
                    cell.alignment = title_alignment
            # Headings and titles
            for idx, entry in enumerate(queryset, start=start_row + 1):
                indent = entry.get('indent', 0)
                self.wsheet.row_dimensions[idx].height = \
                    row_heights[idx - start_row - 1]
                first_cell = self.wsheet.cell(row=idx, column=1)
                first_cell.alignment = first_col_alignments[indent]
                if self.is_practice(entry):
                    first_cell.font = practice_font
                else:
                    first_cell.font = heading_font
                    if not indent:
                        for row_cells in self.wsheet.iter_rows(
                                min_row=idx, max_row=idx):
                            for cell in row_cells:
                                cell.fill = tile_background
            self._report_queries("applied styles on first column in sheet '%s'"
                % title)

            if self.add_expanded_styles:
                headers = self.get_headings(renderer_context=renderer_context)
                for row_cells in self.wsheet.iter_rows(
                        min_row=start_row + 1, max_row=self.wsheet.max_row):
                    first = True
                    for idx, cell in enumerate(row_cells):
                        if first:
                            first = False
                            continue
                        cell.alignment = inner_cell_alignment
                        if (intrinsic_value_headers and
                            headers[idx] in intrinsic_value_headers):
                            try:
                                value = int(cell.value)
                                if value:
                                    cell.border = bordered
                                    cell.fill = self.get_value_fill(value)
                            except (TypeError, ValueError):
                                pass
                self._report_queries("applied all styles in sheet '%s'" % title)

        self._report_queries("written sheet '%s'" % title)


    def render(self, data, accepted_media_type=None, renderer_context=None):
        self._start_time()
        # We need to run `get_queryset` before `get_headings` so we know
        # how many columns to display for implementation rate.
        self.write_sheet(queryset=data.get('results', []), title="Practices",
            renderer_context=renderer_context)

        return self.flush_writer(self.wbook)

    def optimal_cell_sizes(self):
        #pylint:disable=too-many-locals
        iter_rows = self.wsheet.iter_rows()
        if iter_rows != tuple():
            col_width = []

            for i in range(len(next(iter_rows))):
                col_letter = get_column_letter(i + 1)

                minimum_width = 10.0
                current_width = self.wsheet.column_dimensions[col_letter].width
                if not current_width or current_width < minimum_width:
                    self.wsheet.column_dimensions[col_letter].width = \
                        minimum_width
                col_width.append(
                    self.wsheet.column_dimensions[col_letter].width)

            for i, row in enumerate(self.wsheet):
                default_height = 12.5

                multiples_of_font_size = [default_height]
                for j, cell in enumerate(row):
                    if cell.value is not None:
                        mul = 0
                        for val in str(cell.value).split('\n'):
                            mul += math.ceil(
                                len(val) / col_width[j]) * cell.font.size

                        if mul > 0:
                            multiples_of_font_size.append(mul)

                original_height = self.wsheet.row_dimensions[i+1].height
                if original_height is None:
                    original_height = default_height

                new_height = max(multiples_of_font_size)
                max_height = default_height * 10
                if original_height < new_height:
                    if new_height < max_height:
                        self.wsheet.row_dimensions[i + 1].height = new_height
                    else:
                        self.wsheet.row_dimensions[i + 1].height = max_height
