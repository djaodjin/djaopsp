# Copyright (c) 2026, DjaoDjin inc.
# see LICENSE.
import io, math

from django.http import HttpResponse
from django.views.generic import ListView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter
from pages.api.elements import PageElementAPIView
from survey.helpers import datetime_or_now
from survey.mixins import TimersMixin
from survey.models import Unit

from .base import PracticesXLSXRenderer
from ..compat import gettext_lazy as _
from ..helpers import as_valid_sheet_title


class ContentDetailDownloadView(PageElementAPIView):
    """
    Download the practices as a spreadsheet
    """
    schema = None
    base_headers = ['']
    intrinsic_value_headers = ['Environmental', 'Ops/maintenance',
        'Financial', 'Implementation ease', 'AVERAGE VALUE']

    renderer_classes = [PracticesXLSXRenderer]

    def finalize_response(self, request, response, *args, **kwargs):
        resp = super(ContentDetailDownloadView, self).finalize_response(
            request, response, *args, **kwargs)
        filename = datetime_or_now().strftime(
            self.element.slug + '-%Y%m%d.xlsx')
        resp['Content-Disposition'] = 'attachment; filename="{}"'.format(
            filename)
        return resp


class PracticesSpreadsheetView(TimersMixin, ListView):

    add_style = True
    add_expanded_styles = True
    depth = 0
    basename = 'practices'
    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    base_headers = ['']
    peer_value_headers = []
    intrinsic_value_headers = []

    valueFills = [
        PatternFill(fill_type=FILL_SOLID, fgColor='FF9CD76B'), # green-level-0
        PatternFill(fill_type=FILL_SOLID, fgColor='FF69B02B'), # green-level-1
        PatternFill(fill_type=FILL_SOLID, fgColor='FF007C3F'), # green-level-2
        PatternFill(fill_type=FILL_SOLID, fgColor='FFFFD700'), # green-level-3
    ]

    def __init__(self, *args):
        super(PracticesSpreadsheetView, self).__init__(*args)
        self.wbook = None
        self.wsheet = None

    # Methods to be redefined in subclasses
    def get_title(self):
        """
        Returns title with source information for the data on the sheet
        """
        return []

    def get_headings(self):
        """
        Returns column headings for the data on the sheet
        """
        return []

    def format_row(self, entry, key=None):
        """
        Returns a formatted row of data to write on the sheet
        """
        #pylint:disable=unused-argument
        return []

    def get_queryset(self):
        """
        Returns data from the database
        """
        return []

    @staticmethod
    def _get_row_header(entry):
        default_unit = entry.get('default_unit', {})
        default_unit_choices = []
        if default_unit:
            try:
                default_unit = default_unit.slug
            except AttributeError:
                default_unit_choices = default_unit.get('choices', [])
                default_unit = default_unit.get('slug', "")

        title = entry.get('title', "")
        ref_num = entry.get('ref_num')
        if ref_num:
            title = "%s %s" % (ref_num, title)

        default_unit = entry.get('default_unit', {})
        if default_unit:
            try:
                default_unit_system = default_unit.system
                default_unit_title = default_unit.title
            except AttributeError:
                default_unit_system = default_unit.get('system')
                default_unit_title = default_unit.get('title')
            if default_unit_system in Unit.METRIC_SYSTEMS:
                title += " (in %s)" % default_unit_title
            else:
                subtitle = ""
                try:
                    default_unit_choices = default_unit.choices
                except AttributeError:
                    default_unit_choices = default_unit.get('choices', [])
                for choice in default_unit_choices:
                    text = choice.get('text', "").strip()
                    descr = choice.get('descr', "").strip()
                    if text != descr:
                        subtitle += "\n%s - %s" % (text, descr)
                if subtitle:
                    title += "\n" + subtitle

        return title


    def get_value_fill(self, val):
        idx = val - 1
        if idx < len(self.valueFills):
            return self.valueFills[idx]
        return self.valueFills[-1]

    def is_practice(self, entry):
        return 'default_unit' in entry and entry['default_unit']

    def create_writer(self, title=None):
        #pylint:disable=unused-argument
        if not hasattr(self, 'wbook') or not self.wbook:
            self.wbook = Workbook()
            self.wsheet = self.wbook.active
            if title:
                self.wsheet.title = as_valid_sheet_title(title)
        else:
            self.wsheet = self.wbook.create_sheet(
                as_valid_sheet_title(title))

    def flush_writer(self):
        # XXX Running `optimal_cell_sizes` takes about a minute
        # on large datasets. disabled for now.
        # self.optimal_cell_sizes()
        self._report_queries("optimal cell sizes computed")
        # Write out the Excel file. That still takes an inordinate
        # amount of time for some reason, but we have no choice.
        content = io.BytesIO()
        self.wbook.save(content)
        self._report_queries("workbook content saved")
        content.seek(0)
        return content

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access,unused-argument
        self.wsheet.append(row)

    def write_headers(self):
        """
        Write table headers in the worksheet
        """
        self.writerow(self.get_title())
        super_headers = []
        nb_peer_value_headers = 0
        nb_intrinsic_value_headers = 0
        headers = self.get_headings() # We have to call `get_headings`
        # before retrieving `peer_value_headers`.
        if self.peer_value_headers:
            nb_peer_value_headers = 1
            super_headers += ['Peer-based value'] # for correct nb cols
            for header in self.peer_value_headers:
                super_headers += ["" for unused in range(0, len(header[1]))]
                nb_peer_value_headers += len(header[1])
        if self.intrinsic_value_headers:
            nb_intrinsic_value_headers = len(self.intrinsic_value_headers)
            super_headers += ['Intrinsic value'] + [
                "" for unused in range(1, nb_intrinsic_value_headers)]
        if super_headers:
            super_headers = ([
                "" for unused in range(0, len(self.base_headers))] +
                super_headers)
            self.writerow(super_headers)
            if nb_peer_value_headers:
                first_col = chr(ord('A') + len(self.base_headers))
                last_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers - 1)
                self.wsheet.merge_cells('%s1:%s1' % (first_col, last_col))
            if nb_intrinsic_value_headers:
                first_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers)
                last_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers + nb_intrinsic_value_headers - 1)
                self.wsheet.merge_cells('%s1:%s1' % (first_col, last_col))
        self.writerow(headers)

    def write_sheet(self, title="", key=None, queryset=None):
        #pylint:disable=too-many-locals,too-many-nested-blocks
        if not queryset:
            queryset = self.get_queryset()

        self.create_writer(title=title)
        self.write_headers()
        self._report_queries(
            "headers written in sheet '%s'" % title)
        start_row = self.wsheet.max_row
        for entry in queryset:
            row = self.format_row(entry, key=key)
            self.writerow(row)
        self._report_queries("rows written in sheet '%s'" % title)

        if self.add_style:
            max_indent = 0
            height_ratio = 2.8
            mm_to_pts_ratio = 0.5102
            max_heading_width = 132.15 * mm_to_pts_ratio
            row_heights = []
            for idx, entry in enumerate(queryset):
                max_indent = max(max_indent, entry.get('indent', 0))
                #max_heading_width = max(max_heading_width,
                #    entry['indent'] + len(entry['title']))
                nb_wrapped_lines = max(int((
                    entry['indent'] + len(entry['title']))
                    * 0.9 / max_heading_width) + 1, 1)
                row_heights += [nb_wrapped_lines * 5.29 * height_ratio]

            title_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            heading_font = Font(
                name='Calibri', size=12, bold=False, italic=False,
                vertAlign='baseline', underline='none', strike=False,
                color='54BAD8')
            practice_font = Font(
                name='Calibri', size=12, bold=False, italic=False,
                vertAlign='baseline', underline='none', strike=False,
                color='777777')
            tile_background = PatternFill(
                fill_type=FILL_SOLID, fgColor='FFFFA6')
            bordered = Border(
                left=Side(border_style=BORDER_THIN, color='FF000000'),
                right=Side(border_style=BORDER_THIN, color='FF000000'),
                top=Side(border_style=BORDER_THIN, color='FF000000'),
                bottom=Side(border_style=BORDER_THIN, color='FF000000'))
            inner_cell_alignment = Alignment(
                horizontal="left", vertical='top', wrap_text=True)
            # Creates a table of indentation for row titles
            first_col_alignments = [Alignment(horizontal='left', vertical='top',
                            indent=indent, wrap_text=True)
                for indent in range(0, max_indent + 1)]

            self._report_queries("computed styles for sheet '%s'" % title)

            # Column headers
            for column in self.wsheet.columns:
                try:
                    column_letter = column[0].column_letter
                    self.wsheet.column_dimensions[column_letter].width = \
                        22.58 * mm_to_pts_ratio
                except AttributeError:
                    continue
            self.wsheet.column_dimensions['A'].width = max_heading_width
            for row_cells in self.wsheet.iter_rows(
                    min_row=0, max_row=start_row):
                for cell in row_cells:
                    cell.alignment = title_alignment
            # Headings and titles
            for idx, entry in enumerate(queryset, start=start_row + 1):
                indent = entry.get('indent', 0)
                self.wsheet.row_dimensions[idx].height = \
                    row_heights[idx - start_row - 1]
                first_cell = self.wsheet.cell(row=idx, column=1)
                first_cell.alignment = first_col_alignments[indent]
                if self.is_practice(entry):
                    first_cell.font = practice_font
                else:
                    first_cell.font = heading_font
                    if not indent:
                        for row_cells in self.wsheet.iter_rows(
                                min_row=idx, max_row=idx):
                            for cell in row_cells:
                                cell.fill = tile_background
            self._report_queries("applied styles on first column in sheet '%s'"
                % title)

            if self.add_expanded_styles:
                headers = self.get_headings()
                for row_cells in self.wsheet.iter_rows(
                        min_row=start_row + 1, max_row=self.wsheet.max_row):
                    first = True
                    for idx, cell in enumerate(row_cells):
                        if first:
                            first = False
                            continue
                        cell.alignment = inner_cell_alignment
                        if (self.intrinsic_value_headers and
                            headers[idx] in self.intrinsic_value_headers):
                            try:
                                value = int(cell.value)
                                if value:
                                    cell.border = bordered
                                    cell.fill = self.get_value_fill(value)
                            except (TypeError, ValueError):
                                pass
                self._report_queries("applied all styles in sheet '%s'" % title)

        self._report_queries("written sheet '%s'" % title)


    def get(self, request, *args, **kwargs):
        #pylint: disable=unused-argument,too-many-locals
        self._start_time()

        # We need to run `get_queryset` before `get_headings` so we know
        # how many columns to display for implementation rate.
        self.write_sheet(title="Practices", queryset=self.get_queryset())

        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = \
            'attachment; filename="{}"'.format(self.get_filename())
        return resp

    def optimal_cell_sizes(self):
        #pylint:disable=too-many-locals
        iter_rows = self.wsheet.iter_rows()
        if iter_rows != tuple():
            col_width = []

            for i in range(len(next(iter_rows))):
                col_letter = get_column_letter(i + 1)

                minimum_width = 10.0
                current_width = self.wsheet.column_dimensions[col_letter].width
                if not current_width or current_width < minimum_width:
                    self.wsheet.column_dimensions[col_letter].width = \
                        minimum_width
                col_width.append(
                    self.wsheet.column_dimensions[col_letter].width)

            for i, row in enumerate(self.wsheet):
                default_height = 12.5

                multiples_of_font_size = [default_height]
                for j, cell in enumerate(row):
                    if cell.value is not None:
                        mul = 0
                        for val in str(cell.value).split('\n'):
                            mul += math.ceil(
                                len(val) / col_width[j]) * cell.font.size

                        if mul > 0:
                            multiples_of_font_size.append(mul)

                original_height = self.wsheet.row_dimensions[i+1].height
                if original_height is None:
                    original_height = default_height

                new_height = max(multiples_of_font_size)
                max_height = default_height * 10
                if original_height < new_height:
                    if new_height < max_height:
                        self.wsheet.row_dimensions[i + 1].height = new_height
                    else:
                        self.wsheet.row_dimensions[i + 1].height = max_height

    def get_filename(self):
        basename = self.basename
        if hasattr(self, 'account'):
            basename = '-'.join([str(self.account), self.basename])
        return datetime_or_now().strftime(basename + '-%Y%m%d.xlsx')
