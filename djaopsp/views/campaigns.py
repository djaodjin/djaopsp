# Copyright (c) 2022, DjaoDjin inc.
# see LICENSE.

"""
Views used to edit assessment campaigns
"""

import io, math

from deployutils.helpers import datetime_or_now, update_context_urls
from django.http import HttpResponse
from django.views.generic import ListView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter
from pages.views.elements import PageElementEditableView
from survey.mixins import CampaignMixin

from ..api.campaigns import CampaignContentMixin
from ..compat import reverse
from ..helpers import as_valid_sheet_title


class CampaignEditView(CampaignMixin, PageElementEditableView):

    def get_template_names(self):
        if self.is_prefix:
            # It is not a leaf, let's return the list view
            return ['survey/campaigns/campaign.html']
        return super(CampaignEditView, self).get_template_names()

    def get_context_data(self, **kwargs):
        context = super(CampaignEditView, self).get_context_data(**kwargs)

        campaign_slug = self.kwargs.get('campaign')
        context.update({'campaign': self.campaign})
        urls = {
            'api_editable_segments': reverse('api_campaign_editable_segments',
                args=(self.account, campaign_slug))
        }
        if self.is_prefix:
            # Editor
            urls.update({
                'api_practice_typeahead': reverse(
                    'pages_api_editables_index', args=(self.account,)),
                'api_content': reverse('api_campaign_editable_content',
                        args=(self.account, campaign_slug)),
                'campaign_download': reverse('campaign_download',
                    args=(self.account, campaign_slug)),
                'api_alias_node': reverse('pages_api_alias_node', args=(
                    self.account, '')),
                'api_mirror_node': reverse('pages_api_mirror_node',
                    args=(self.account, '')),
                'api_move_node': reverse('pages_api_move_node',
                    args=(self.account, '')),
            })
        update_context_urls(context, urls)
        return context


class CampaignXLSXView(CampaignContentMixin, ListView):

    depth = 0
    basename = 'practices'
    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def create_writer(self, headings, title=None):
        #pylint:disable=unused-argument
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            self.wsheet.title = as_valid_sheet_title(title)
        self.heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF0071BB')
        self.heading_alignment = Alignment(wrap_text=True)
        self.border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000'))
        self.text_center = Alignment(horizontal='center')

    def flush_writer(self):
        self.optimal_cell_sizes()
        # Write out the Excel file.
        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access,unused-argument
        self.wsheet.append(row)

    def get(self, request, *args, **kwargs):
        #pylint: disable=unused-argument,too-many-locals
        segments = [seg for seg in self.segments_available if seg['path']]
        headings = [''] + [seg['title'] for seg in segments]
        self.create_writer(headings, title="Practices")
        self.writerow(headings)

        title_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        for row_cells in self.wsheet.iter_rows(
                min_row=self.wsheet.max_row,
                max_row=self.wsheet.max_row):
            for cell in row_cells:
                cell.alignment = title_alignment

        heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='54BAD8')
        practice_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='777777')
        tile_background = PatternFill(fill_type=FILL_SOLID, fgColor='FFFFA6')
        inner_cell_alignment = Alignment(
            horizontal="left",
            vertical='top',
            wrap_text=True)
        heading_width = 3.32

        results = self.get_queryset()
        for entry in results:
            row = [entry['title']]
            for seg in segments:
                tags = entry.get('extra', {}).get('tags', [])
                if seg['path'] in tags:
                    row += [entry['title']]
                else:
                    row += ['']
            self.writerow(row)
            self.wsheet.cell(row=self.wsheet.max_row, column=1).alignment = \
                Alignment(horizontal='left', vertical='top',
                    indent=entry['indent'],
                    wrap_text=True)
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.wsheet.max_row,
                    max_row=self.wsheet.max_row):
                first = True
                for cell in row_cells:
                    if not first:
                        cell.alignment = inner_cell_alignment
                    if 'url' in entry and entry['url']:
                        cell.font = practice_font
                    else:
                        cell.font = heading_font
                        heading_width = max(heading_width,
                            entry['indent'] + len(entry['title']))
                    if entry['indent'] == 0:
                        cell.fill = tile_background
                    first = False

        self.wsheet.column_dimensions['A'].width = 0.332 * heading_width

        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = \
            'attachment; filename="{}"'.format(self.get_filename())
        return resp

    def optimal_cell_sizes(self):
        #pylint:disable=too-many-locals
        iter_rows = self.wsheet.iter_rows()
        if iter_rows != tuple():
            col_width = []

            for i in range(len(next(iter_rows))):
                col_letter = get_column_letter(i + 1)

                minimum_width = 10.0
                current_width = self.wsheet.column_dimensions[col_letter].width
                if not current_width or current_width < minimum_width:
                    self.wsheet.column_dimensions[col_letter].width = \
                        minimum_width
                col_width.append(
                    self.wsheet.column_dimensions[col_letter].width)

            for i, row in enumerate(self.wsheet):
                default_height = 12.5

                multiples_of_font_size = [default_height]
                for j, cell in enumerate(row):
                    if cell.value is not None:
                        mul = 0
                        for val in str(cell.value).split('\n'):
                            mul += math.ceil(
                                len(val) / col_width[j]) * cell.font.size

                        if mul > 0:
                            multiples_of_font_size.append(mul)

                original_height = self.wsheet.row_dimensions[i+1].height
                if original_height is None:
                    original_height = default_height

                new_height = max(multiples_of_font_size)
                max_height = default_height * 10
                if original_height < new_height:
                    if new_height < max_height:
                        self.wsheet.row_dimensions[i + 1].height = new_height
                    else:
                        self.wsheet.row_dimensions[i + 1].height = max_height

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')
