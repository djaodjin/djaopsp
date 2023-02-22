# Copyright (c) 2022, DjaoDjin inc.
# see LICENSE.
from __future__ import unicode_literals

import io, math

from deployutils.helpers import datetime_or_now
from django.http import HttpResponse
from django.views.generic import ListView
from extended_templates.backends.pdf import PdfTemplateResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter
from pages.models import PageElement

from ..api.samples import AssessmentContentMixin
from ..helpers import as_valid_sheet_title


class PracticesSpreadsheetView(ListView):

    depth = 0
    basename = 'practices'
    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    base_headers = ['']
    peer_value_headers = []
    intrinsic_value_headers = []

    valueFills = [
        PatternFill(fill_type=FILL_SOLID, fgColor='FF9CD76B'), # green-level-0
        PatternFill(fill_type=FILL_SOLID, fgColor='FF69B02B'), # green-level-1
        PatternFill(fill_type=FILL_SOLID, fgColor='FF007C3F'), # green-level-2
        PatternFill(fill_type=FILL_SOLID, fgColor='FFFFD700'), # green-level-3
    ]

    def get_title_row(self):
        return []

    def get_value_fill(self, val):
        idx = val - 1
        if idx < len(self.valueFills):
            return self.valueFills[idx]
        return self.valueFills[-1]

    # Methods to be redefined in subclasses
    def get_headings(self):
        return []

    def format_row(self, entry):
        return []

    def get_queryset(self):
        return []

    def is_practice(self, entry):
        return 'default_unit' in entry and entry['default_unit']

    def create_writer(self, title=None):
        #pylint:disable=unused-argument
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            self.wsheet.title = as_valid_sheet_title(title)
        self.heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF0071BB')
        self.heading_alignment = Alignment(wrap_text=True)
        self.border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000'))
        self.text_center = Alignment(horizontal='center')

    def flush_writer(self):
        self.optimal_cell_sizes()
        # Write out the Excel file.
        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access,unused-argument
        self.wsheet.append(row)

    def get(self, request, *args, **kwargs):
        #pylint: disable=unused-argument,too-many-locals

        # We need to run `get_queryset` before `get_headings` so we know
        # how many columns to display for implementation rate.
        queryset = self.get_queryset()
        headers = self.get_headings()
        self.create_writer(title="Practices")
        self.writerow(self.get_title_row())
        super_headers = []
        nb_peer_value_headers = 0
        nb_intrinsic_value_headers = 0
        if self.peer_value_headers:
            nb_peer_value_headers = 1
            super_headers += ['Peer-based value'] # for correct nb cols
            for header in self.peer_value_headers:
                super_headers += ["" for unused in range(0, len(header[1]))]
                nb_peer_value_headers += len(header[1])
        if self.intrinsic_value_headers:
            nb_intrinsic_value_headers = len(self.intrinsic_value_headers)
            super_headers += ['Intrinsic value'] + [
                "" for unused in range(1, nb_intrinsic_value_headers)]
        if super_headers:
            super_headers = ([
                "" for unused in range(0, len(self.base_headers))] +
                super_headers)
            self.writerow(super_headers)
            if nb_peer_value_headers:
                first_col = chr(ord('A') + len(self.base_headers))
                last_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers - 1)
                self.wsheet.merge_cells('%s1:%s1' % (first_col, last_col))
            if nb_intrinsic_value_headers:
                first_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers)
                last_col = chr(ord('A') + len(self.base_headers) +
                    nb_peer_value_headers + nb_intrinsic_value_headers - 1)
                self.wsheet.merge_cells('%s1:%s1' % (first_col, last_col))
        self.writerow(headers)

        title_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        for row_cells in self.wsheet.iter_rows(
                min_row=self.wsheet.max_row,
                max_row=self.wsheet.max_row):
            for cell in row_cells:
                cell.alignment = title_alignment

        heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='54BAD8')
        practice_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='777777')
        tile_background = PatternFill(fill_type=FILL_SOLID, fgColor='FFFFA6')
        inner_cell_alignment = Alignment(
            horizontal="left",
            vertical='top',
            wrap_text=True)
        heading_width = 3.32

        for entry in queryset:
            row = self.format_row(entry)
            self.writerow(row)
            self.wsheet.cell(row=self.wsheet.max_row, column=1).alignment = \
                Alignment(horizontal='left', vertical='top',
                    indent=entry['indent'],
                    wrap_text=True)
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.wsheet.max_row,
                    max_row=self.wsheet.max_row):
                first = True
                for idx, cell in enumerate(row_cells):
                    if not first:
                        cell.alignment = inner_cell_alignment
                    if self.is_practice(entry):
                        cell.font = practice_font
                    else:
                        cell.font = heading_font
                        heading_width = max(heading_width,
                            entry['indent'] + len(entry['title']))
                    if entry['indent'] == 0:
                        cell.fill = tile_background
                    if headers[idx] in self.intrinsic_value_headers:
                        try:
                            value = int(cell.value)
                            if value:
                                cell.border = self.border
                                cell.fill = self.get_value_fill(value)
                        except (TypeError, ValueError):
                            pass
                    first = False

        self.wsheet.column_dimensions['A'].width = 0.332 * heading_width

        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = \
            'attachment; filename="{}"'.format(self.get_filename())
        return resp

    def optimal_cell_sizes(self):
        #pylint:disable=too-many-locals
        iter_rows = self.wsheet.iter_rows()
        if iter_rows != tuple():
            col_width = []

            for i in range(len(next(iter_rows))):
                col_letter = get_column_letter(i + 1)

                minimum_width = 10.0
                current_width = self.wsheet.column_dimensions[col_letter].width
                if not current_width or current_width < minimum_width:
                    self.wsheet.column_dimensions[col_letter].width = \
                        minimum_width
                col_width.append(
                    self.wsheet.column_dimensions[col_letter].width)

            for i, row in enumerate(self.wsheet):
                default_height = 12.5

                multiples_of_font_size = [default_height]
                for j, cell in enumerate(row):
                    if cell.value is not None:
                        mul = 0
                        for val in str(cell.value).split('\n'):
                            mul += math.ceil(
                                len(val) / col_width[j]) * cell.font.size

                        if mul > 0:
                            multiples_of_font_size.append(mul)

                original_height = self.wsheet.row_dimensions[i+1].height
                if original_height is None:
                    original_height = default_height

                new_height = max(multiples_of_font_size)
                max_height = default_height * 10
                if original_height < new_height:
                    if new_height < max_height:
                        self.wsheet.row_dimensions[i + 1].height = new_height
                    else:
                        self.wsheet.row_dimensions[i + 1].height = max_height

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')


class ImproveContentPDFView(AssessmentContentMixin, ListView):

    http_method_names = ['get']
    template_name = 'app/prints/improve.html'
    content_type = 'application/pdf'
    indent_step = '    '

    def __init__(self, **kwargs):
        super(ImproveContentPDFView, self).__init__(**kwargs)
        self.table_of_content = []

    def get(self, request, *args, **kwargs):
        self.object_list = []
        for item in self.get_queryset():
            if 'default_unit' in item:
                element_text_qs = PageElement.objects.filter(
                    slug=item['slug']).values_list('text', flat=True)
                element_text = element_text_qs[0] if element_text_qs else ""
                item.update({'text': element_text})
                self.object_list += [item]
            self.table_of_content += [item]
        context = self.get_context_data(**kwargs)
        context.update({'table_of_content': self.table_of_content})
        return PdfTemplateResponse(request, self.template_name, context)
