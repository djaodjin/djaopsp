# Copyright (c) 2019, DjaoDjin inc.
# see LICENSE.
from __future__ import unicode_literals

import csv, logging, io
from collections import OrderedDict

from deployutils.apps.django.templatetags.deployutils_prefixtags import (
    site_prefixed)
from deployutils.helpers import datetime_or_now, update_context_urls
from django.http import HttpResponse
from django.views.generic.list import ListView
from django.utils import six
from extended_templates.backends.pdf import PdfTemplateResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID
from pages.models import PageElement

from .assessments import AssessmentBaseMixin, AssessmentView
from .benchmark import PrintableChartsMixin
from ..compat import reverse
from ..mixins import ContentCut, ImprovementQuerySetMixin
from ..helpers import as_valid_sheet_title
from ..templatetags.navactive import (
    assessment_choices as assessment_choices_base)


LOGGER = logging.getLogger(__name__)


class ImprovementView(ImprovementQuerySetMixin, AssessmentView):

    template_name = 'envconnect/improve.html'
    breadcrumb_url = 'improve'

    def get_breadcrumb_url(self, path):
        organization = self.kwargs.get('organization', None)
        if organization:
            return reverse(
                'improve_organization_redirect', args=(organization, path))
        return super(ImprovementView, self).get_breadcrumb_url(path)

    def get_report_tree(self, node=None, from_root=None, cut=ContentCut(),
                        load_text=False):
        self.get_or_create_improve_sample()
        root = super(ImprovementView, self).get_report_tree(
            node=node, from_root=from_root, cut=cut, load_text=load_text)
        if root:
            self.decorate_with_breadcrumbs(root)
        # Removes framework additional questions and other assessments
        # questions which are irrelevant when making an improvement plan.
        to_remove = []
        for key, child_node in six.iteritems(root[1]):
            tags = child_node[0].get('tag', "")
            if tags and 'metrics' in tags:
                to_remove += [key]
        for key in to_remove:
            del root[1][key]
        return root

    def get_context_data(self, **kwargs):
        context = super(ImprovementView, self).get_context_data(**kwargs)
        segment_url, segment_prefix, segment_element = self.segment

        organization = context['organization']
        update_context_urls(context, {
            'api_improvements': reverse('api_improvement_base',
                args=(organization,)),
            'api_organizations': site_prefixed("/api/profile/"),
            'api_viewers': site_prefixed(
                "/api/profile/%s/roles/viewers/" % self.account),
        })
        if self.improvement_sample:
            update_context_urls(context, {
                'print': reverse('improve_organization_sample_print',
                    args=(organization, self.improvement_sample,
                        '/%s' % segment_prefix.split('/')[-1])),
                'download': reverse('improve_organization_sample_download',
                    args=(organization, self.improvement_sample,
                        '/%s' % segment_prefix.split('/')[-1]))
            })
        if self.assessment_sample:
            update_context_urls(context, {
                'api_account_benchmark': reverse('api_benchmark',
                args=(organization, self.assessment_sample, segment_prefix))})
        return context


class ImprovementOnlyMixin(ImprovementQuerySetMixin, AssessmentBaseMixin):

    def _improvements_only(self, root=None):
        """
        returns a tree where only nodes leading to leafs which are in queryset.
        """
        if not root:
            # All assessment questions for a segment, regardless
            # of the actual from_path.
            # XXX if we do that, we shouldn't use from_root (i.e. system pages)
            _, trail = self.breadcrumbs
            trail_head = ("/" + trail[0][0].slug.decode('utf-8')
                if six.PY2 else trail[0][0].slug)
            from_trail_head = "/" + "/".join([
                element.slug.decode('utf-8') if six.PY2 else element.slug
                for element in self.get_full_element_path(trail_head)])
            # We use cut=None here so we print out the full assessment
            root = self.get_report_tree(trail[0][0], from_trail_head, cut=None)

        filtered = OrderedDict()
        consumption = root[0].get('consumption', {})
        if consumption:
            planned = consumption.get('planned', None)
            if planned:
                page_element = PageElement.objects.get(slug=root[0].get('slug'))
                root[0]['consumption'].update({
                    'title': page_element.title,
                    'text': page_element.text
                })
                return (root[0], filtered)
        for path, nodes in six.iteritems(root[1]):
            filtered_nodes = self._improvements_only(nodes)
            if filtered_nodes:
                filtered.update({path: filtered_nodes})
        if filtered:
            return (root[0], filtered)
        return None


class ImprovementSpreadsheetView(ImprovementOnlyMixin, ListView):

    basename = 'improvements'
    fixed_headings = ['Practice', 'Assessment', 'Planned']
    value_headings = [
        'Environmental', 'Ops/maintenance', 'Financial',
        'Implementation ease', 'AVERAGE VALUE']
    indent_step = '    '

    @property
    def assessment_choices(self):
        if not hasattr(self, '_assessment_choices'):
            segment_url, segment_prefix, segment_element = self.segment
            if segment_element.slug == 'framework':
                self._assessment_choices = [str(choice)
                    for choice in assessment_choices_base(segment_element.slug)]
            else:
                self._assessment_choices = [
                    'Implementation rate', 'Opportunity score']
        return self._assessment_choices

    def write_tree(self, root, indent=''):
        if not root:
            return
        element = root[0].get('consumption', None)
        if element:
            # Use account contextual opportunity instead of base opportunity
            # in `element['opportunity']`
            opportunity = root[0].get('accounts', {}).get(
                self.account.pk, {}).get('opportunity_numerator')
            planned = element.get('planned')
            if planned == element['implemented']:
                planned = ""
            # We reached a leaf
            segment_url, segment_prefix, segment_element = self.segment
            if segment_element.slug == 'framework':
                self.writerow([
                    indent + element['title'],
                    element['implemented'], planned] + [
                        element['rate'].get(choice, "")
                        for choice in self.assessment_choices
                ], leaf=True)
            else:
                self.writerow([
                    indent + element['title'],
                    element['implemented'],
                    "yes" if planned else "",
                    element['rate'],
                    "%.2f" % opportunity,
                    element['environmental_value'],
                    element['business_value'],
                    element['implementation_ease'],
                    element['profitability'],
                    element['avg_value']
                ], leaf=True)
        else:
            element = root[0]
            self.writerow([indent + element['title']])
            for _, nodes in six.iteritems(root[1]):
                self.write_tree(nodes, indent=indent + self.indent_step)

    def get(self, request, *args, **kwargs): #pylint: disable=unused-argument
        root = self._improvements_only()
        segment_url, segment_prefix, segment_element = self.segment
        segment_title = segment_element.title
        self.create_writer(title=segment_title)
        self.writerow(["%s - Environmental practices - Improvement plan" %
            self.account.printable_name], leaf=True)
        self.writerow(self.fixed_headings, leaf=True)
        self.writerow(self.get_headings(), leaf=True)
        self.write_tree(root)
        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = \
            'attachment; filename="{}"'.format(self.get_filename())
        return resp

    def get_headings(self):
        # Technically should be a default_metric on questions but right now
        # it is a tag in the icon tag (one level below the segment).
        segment_url, segment_prefix, segment_element = self.segment
        if segment_element.slug == 'framework':
            return self.fixed_headings + self.assessment_choices
        return self.fixed_headings + self.assessment_choices + self.value_headings

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.csv')


class ImprovementCSVView(ImprovementSpreadsheetView):

    content_type = 'text/csv'

    def writerow(self, row, leaf=False):
        #pylint:disable=unused-argument
        self.csv_writer.writerow([
            rec.encode('utf-8') if six.PY2 else rec
            for rec in row])

    def create_writer(self, title=None):
        #pylint:disable=unused-argument
        self.content = io.StringIO()
        self.csv_writer = csv.writer(self.content)

    def flush_writer(self):
        self.content.seek(0)
        return self.content

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.csv')


class ImprovementXLSXView(PrintableChartsMixin, ImprovementSpreadsheetView):

    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    valueFills = [
        PatternFill(fill_type=FILL_SOLID, fgColor='FF9CD76B'), # green-level-0
        PatternFill(fill_type=FILL_SOLID, fgColor='FF69B02B'), # green-level-1
        PatternFill(fill_type=FILL_SOLID, fgColor='FF007C3F'), # green-level-2
        PatternFill(fill_type=FILL_SOLID, fgColor='FFFFD700'), # green-level-3
    ]

    def create_writer(self, title=None):
        col_scale = 11.9742857142857
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            self.wsheet.title = as_valid_sheet_title(title)
        self.wsheet.row_dimensions[1].height = 0.36 * (6 * col_scale)
        self.wsheet.column_dimensions['A'].width = 6.56 * col_scale
        for col_num in range(1,
                len(self.fixed_headings + self.assessment_choices)):
            self.wsheet.column_dimensions[chr(ord('A') + col_num)].width = \
                0.99 * col_scale
        for col_num in range(
                len(self.fixed_headings + self.assessment_choices),
                len(self.get_headings())):
            self.wsheet.column_dimensions[chr(ord('A') + col_num)].width = \
                0.99 * col_scale
        self.heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF0071BB')
        self.heading_alignment = Alignment(wrap_text=True)
        self.border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000'))
        self.text_center = Alignment(horizontal='center')

    def flush_writer(self):
        #pylint:disable=protected-access,too-many-statements
        bold_font = Font(
            name='Calibri', size=11, bold=True, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        self.wsheet.append([])
        self.wsheet.append(["Assessment"])
        self.wsheet.row_dimensions[self.wsheet._current_row].font = bold_font
        self.wsheet.append(["    Extent to which you"\
            " indicated the practice is implemented across"\
            " activities/services/products/offices/facilities to which"\
            " it could apply."])
        self.wsheet.append(["Planned"])
        self.wsheet.row_dimensions[self.wsheet._current_row].font = bold_font
        self.wsheet.append(["    Extent to which you"\
            " have planned to improve on this practice."])
        self.wsheet.append(["Implementation rate"])
        self.wsheet.row_dimensions[self.wsheet._current_row].font = bold_font
        self.wsheet.append(["    Percentage of peer"\
            " respondents that have implemented a best practice."])
        self.wsheet.append(["Opportunity score"])
        self.wsheet.row_dimensions[self.wsheet._current_row].font = bold_font
        self.wsheet.append(["    Percentage points by which"\
            " your score could increase if this best practice is implemented"\
            " to full extent. See FAQs for scoring methodology"\
            " and calculations."])
        alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=False,
            shrink_to_fit=False, indent=0)
        title_fill = PatternFill(fill_type=FILL_SOLID, fgColor='FFDDD9C5')
        title_font = Font(
            name='Calibri', size=20, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        subtitle_fill = PatternFill(fill_type=FILL_SOLID, fgColor='FFEEECE2')
        subtitle_font = Font(
            name='Calibri', size=10, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        subtitle_alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=True,
            shrink_to_fit=True, indent=0)
        # Implementation Note: We style the cells here instead of the rows
        # otherwise opening the file on Excel leads to weird background coloring
        # (LibreOffice looks fine).
        cell = self.wsheet['A1']
        cell.fill = title_fill
        cell.font = title_font
        cell.border = self.border
        cell.alignment = alignment
        rate_start_col = chr(ord('A') + len(self.fixed_headings))
        rate_end_col = chr(ord('A') + len(self.get_headings()) - 1)
        for col in range(ord('A'), ord(rate_end_col) + 1):
            cell = self.wsheet['%s2' % chr(col)]
            cell.fill = subtitle_fill
            cell.font = subtitle_font
            cell.border = self.border
            cell.alignment = subtitle_alignment
        for col in range(ord(rate_start_col), ord(rate_end_col) + 1):
            cell = self.wsheet['%s3' % chr(col)]
            cell.fill = subtitle_fill
            cell.font = subtitle_font
            cell.border = self.border
            cell.alignment = subtitle_alignment
        self.wsheet.merge_cells('A1:%s1' % rate_end_col) # Title
        self.wsheet.merge_cells('A2:A3')                 # Practice
        self.wsheet.merge_cells('B2:B3')                 # Assessment
        self.wsheet.merge_cells('C2:C3')                 # Planned
        segment_url, segment_prefix, segment_element = self.segment
        if segment_element.slug == 'framework':
            # implementation rate
            self.wsheet['%s2' % rate_start_col].value = 'Implementation rate'
            self.wsheet.merge_cells(
                '%s2:%s2' % (rate_start_col,
                chr(ord(rate_start_col) + len(self.assessment_choices) - 1)))
        else:
            self.wsheet['D2'].value = 'Implementation rate'
            self.wsheet.merge_cells('D2:D3')             # Implementation rate
            self.wsheet['E2'].value = 'Opportunity score'
            self.wsheet.merge_cells('E2:E3')             # Opportunity score
            self.wsheet['F2'].value = 'Value'
            self.wsheet.merge_cells('F2:J2')             # Values

        # Create "Impact of Improvement Plan" worksheet.
        if False: #pylint:disable=using-constant-test
            wsheet = self.wbook.create_sheet(title="Impact of Improvement Plan")
            # XXX get data from actual score.
            chart = {
                'slug': 'totals',
                'highest_normalized_score': 100,
                'avg_normalized_score': 50,
                'normalized_score': 64,
            }
            self.generate_printable_html([chart])
            image_path = chart['image']
            if image_path.startswith('file://'):
                image_path = image_path[7:]
            img = Image(image_path)
#           img.anchor(wsheet['A1']) # XXX img.anchor is not callable in 2.5.4
            wsheet.add_image(img)

        # Write out the Excel file.
        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')

    def get_value_fill(self, val):
        if val < len(self.valueFills):
            return self.valueFills[val]
        return self.valueFills[-1]

    def write_best_practice_content(self, root, indent=''):
        for element in sorted(
                list(root.keys()), key=lambda node: (node.tag, node.pk)):
            # XXX sort won't exactly match the web presentation
            # which uses RelationShip order
            # (see ``BreadcrumbMixin._build_tree``).
            nodes = root[element]
            if 'text' in nodes:
                # We reached a leaf
                title = element.title.replace('*', '').replace('/', '')
                if len(title) > 31:
                    title = title[0:28] + '...'
                wsheet = self.wbook.create_sheet(title=title)
                wsheet['A1'] = nodes['text']
            else:
                self.write_best_practice_content(
                    nodes, indent=indent + self.indent_step)

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access
        self.wsheet.append(row)
        if leaf:
            if len(row) >= len(self.get_headings()):
                for row_cells in self.wsheet.iter_rows(
                        min_row=self.wsheet._current_row,
                        max_row=self.wsheet._current_row):
                    # Implementation rate
                    for cell_idx in range(
                            len(self.fixed_headings),
                            len(self.fixed_headings + self.assessment_choices)):
                        try:
                            row_cells[cell_idx].value = (
                                int(row_cells[cell_idx].value) / 100)
                            row_cells[cell_idx].style = 'Percent'
                        except ValueError:
                            # We might be printing headings.
                            pass
                    for cell_idx in range(
                            len(self.fixed_headings + self.assessment_choices),
                            len(self.get_headings())):
                        # environmental_value, business_value,
                        # implementation_ease, profitability,
                        # avg_value
                        try:
                            row_cells[cell_idx].border = self.border
                            row_cells[cell_idx].fill = self.get_value_fill(
                                int(row_cells[cell_idx].value))
                            row_cells[cell_idx].value = None
                        except ValueError:
                            # might be the header
                            pass
                    row_cells[0].alignment = self.heading_alignment
        else:
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.wsheet._current_row,
                    max_row=self.wsheet._current_row):
                row_cells[0].font = self.heading_font
                row_cells[0].alignment = self.heading_alignment


class ImprovementPDFView(ImprovementOnlyMixin, ListView):

    http_method_names = ['get']
    template_name = 'envconnect/best_practice_pdf.html'
    indent_step = '    '

    def __init__(self, **kwargs):
        super(ImprovementPDFView, self).__init__(**kwargs)
        self.table_of_content = []
        self.report_items = []

    def write_tree(self, root, indent=''):
        if not root:
            return
        self.table_of_content += [(
            len(indent), root[0].get('title', '(no title)'))]
        consumption = root[0].get('consumption', None)
        if consumption:
            breadcrumbs = root[0].get('breadcrumbs', None)
            if breadcrumbs:
                breadcrumbs.pop()
                consumption.update({'breadcrumbs': breadcrumbs})
            self.report_items += [consumption]
        else:
            for _, nodes in six.iteritems(root[1]):
                self.write_tree(nodes, indent=indent + self.indent_step)

    def get(self, request, *args, **kwargs):
        root = self._improvements_only()
        self.report_items = []
        self.write_tree(root)
        self.object_list = self.report_items
        context = self.get_context_data(**kwargs)
        context.update({'table_of_content': self.table_of_content})
        return PdfTemplateResponse(request, self.template_name, context)
