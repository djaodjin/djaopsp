# Copyright (c) 2017, DjaoDjin inc.
# see LICENSE.
from __future__ import unicode_literals

import csv, datetime, json, logging, io

from django.utils import six
from django.core.urlresolvers import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic.base import TemplateView
from deployutils.crypt import JSONEncoder
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID, FILL_NONE
from openpyxl.worksheet.dimensions import ColumnDimension
from survey.models import Answer

from ..api.benchmark import BenchmarkMixin
from ..mixins import BestPracticeMixin
from ..models import Consumption
from ..serializers import ConsumptionSerializer


LOGGER = logging.getLogger(__name__)


class SelfAssessmentBaseView(BenchmarkMixin, BestPracticeMixin, TemplateView):

    def decorate_with_opportunities(self, leafs):
        #pylint: disable=protected-access
        for consumption in Consumption.objects.with_opportunity(
                filter_out_testing=self._get_filter_out_testing()):
            leaf = leafs.get(consumption.path, None)
            if leaf:
                # If the consumption is not in leafs,
                # we have cut out the tree at an icon or heading level.
                leaf[0]['consumption'] = \
                    ConsumptionSerializer().to_representation(consumption)

    def decorate_with_answers(self, leafs):
        # Implementation Note:
        # `decorate_with_answers` must be called after
        # `decorate_with_opportunities` because it will compute the number
        # of additional points available from the consumption opportunity
        # and the answer.
        for path, values in six.iteritems(leafs):
            if values[0]['consumption']:
                try:
                    answer = Answer.objects.get(
                        question__consumption__path=path,
                        response=self.sample)
                    values[0]['consumption']['implemented'] = answer.text
                    # XXX This is not really the opportunity as defined
                    # in `get_opportunities_sql` but rather the number of points
                    # scored based on the yes/no answer as computed
                    # in `get_scored_answers`.
                    if answer.text == Consumption.YES:
                        values[0]['consumption']['opportunity'] *= (3 - 3)
                    elif answer.text == Consumption.NEEDS_MODERATE_IMPROVEMENT:
                        values[0]['consumption']['opportunity'] *= (3 - 2)
                    elif (answer.text
                          == Consumption.NEEDS_SIGNIFICANT_IMPROVEMENT):
                        values[0]['consumption']['opportunity'] *= (3 - 1)
                    elif answer.text == Consumption.NO:
                        values[0]['consumption']['opportunity'] *= (3 - 0)
                    else:
                        # Not Applicable.
                        values[0]['consumption']['opportunity'] = 0
                except Answer.DoesNotExist:
                    # We have cut out the tree at an icon or heading level.
                    values[0]['consumption']['implemented'] = False

    def decorate_with_improvements(self, leafs):
        # Implementation Note:
        # `decorate_with_improvements` must be called after
        # `decorate_with_opportunities` because it relies on
        # values[0]['consumption'] to exist at the time of call.
        for path, values in six.iteritems(leafs):
            if values[0]['consumption']:
                values[0]['consumption']['planned'] \
                    = Answer.objects.filter(
                        question__consumption__path=path,
                        response__extra='is_planned',
                        response__survey__title=self.report_title,
                        response__account=self.account).exists()

    def attach_benchmarks(self, rollup_tree):
        self._start_time()
        self.get_or_create_assessment_sample()
        leafs = self.get_leafs(rollup_tree)
        self._report_queries("[attach_benchmarks] leafs loaded")
        self.decorate_with_opportunities(leafs)
        self.decorate_with_answers(leafs)
        self.decorate_with_improvements(leafs)
        self._report_queries("[attach_benchmarks] leafs populated")


class SelfAssessmentView(SelfAssessmentBaseView):

    template_name = 'envconnect/self_assessment.html'
    breadcrumb_url = 'report'

    def get_breadcrumb_url(self, path):
        organization = self.kwargs.get('organization', None)
        if organization:
            return reverse('report_organization', args=(organization, path))
        return super(SelfAssessmentView, self).get_breadcrumb_url(path)

    def get_context_data(self, **kwargs):
        context = super(SelfAssessmentView, self).get_context_data(**kwargs)
        from_root, trail = self.breadcrumbs
        root = None
        if trail:
            root = self._build_tree(trail[-1][0], from_root)
            self.attach_benchmarks(root)
            context.update({
                'root': root,
                'entries': json.dumps(root, cls=JSONEncoder)
            })
        organization = context['organization']
        context.update({
            'page_icon': self.icon,
            'response': self.sample,
            'survey': self.sample.survey,
            'role': "tab",
            'score_toggle': self.request.GET.get('toggle', False)})
        self.update_context_urls(context, {
            'api_self_assessment_response': reverse(
                'survey_api_response', args=(organization, self.sample)),
        })
        return context

    def get(self, request, *args, **kwargs):
        from_root, _ = self.breadcrumbs
        if not from_root or from_root == "/":
            return HttpResponseRedirect(reverse('homepage'))
        return super(SelfAssessmentView, self).get(request, *args, **kwargs)


class SelfAssessmentSpreadsheetView(SelfAssessmentBaseView):

    basename = 'assessment'
    indent_step = '    '

    @staticmethod
    def _get_consumption(element):
        return element.get('consumption', None)

    @staticmethod
    def _get_tag(element):
        return element.get('tag', "")

    @staticmethod
    def _get_title(element):
        return element.get('title', "")

    def insert_path(self, tree, parts=None):
        if not parts:
            return tree
        if not parts[0] in tree:
            tree[parts[0]] = {}
        return self.insert_path(tree[parts[0]], parts[1:])

    def writerow(self, row, leaf=False):
        pass

    def write_tree(self, root, indent=''):
        """
        The *root* parameter looks like:
        (PageElement, [(PageElement, [...]), (PageElement, [...]), ...])
        """
        if not root[1]:
            # We reached a leaf
            consumption = self._get_consumption(root[0])
            row = [indent + self._get_title(root[0])]
            if consumption:
                for heading in self.get_headings(self._get_tag(root[0])):
                    if consumption.get('implemented', "") == heading:
                        row += ['X']
                    else:
                        row += ['']
            self.writerow(row, leaf=True)
        else:
            self.writerow([indent + self._get_title(root[0])])
            for element in six.itervalues(root[1]):
                self.write_tree(element, indent=indent + self.indent_step)

    def get(self, *args, **kwargs): #pylint: disable=unused-argument
        # All self-assessment questions for an industry, regardless
        # of the actual from_path.
        # XXX if we do that, we shouldn't use from_root (i.e. system pages)
        _, trail = self.breadcrumbs
        trail_head = ("/"
            + trail[0][0].slug.decode('utf-8') if six.PY2 else trail[0][0].slug)
        from_trail_head = "/" + "/".join([
            element.slug.decode('utf-8') if six.PY2 else element.slug
            for element in self.get_full_element_path(trail_head)])
        # We use cut=None here so we print out the full self-assessment
        root = self._build_tree(trail[0][0], from_trail_head, cut=None)
        self.attach_benchmarks(root)

        self.create_writer(
            self.get_headings(self._get_tag(root[0])),
            title=self._get_title(root[0]))
        self.writerow(
            ["Assessment - Environmental practices"], leaf=True)
        self.writerow(
            ["Practice", "Implemented as a standard practice?"], leaf=True)
        self.writerow(
            [""] + self.get_headings(self._get_tag(root[0])), leaf=True)
        indent = ' '
        for nodes in six.itervalues(root[1]):
            self.writerow([indent + self._get_title(nodes[0])])
            for elements in six.itervalues(nodes[1]):
                self.write_tree(elements, indent=indent + self.indent_step)
        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = 'attachment; filename="{}"'.format(
            self.get_filename())
        return resp

    @staticmethod
    def get_headings(tag):
        return list(Consumption.ASSESSMENT_CHOICES.get(tag,
            Consumption.ASSESSMENT_CHOICES.get('default')))


class SelfAssessmentCSVView(SelfAssessmentSpreadsheetView):

    content_type = 'text/csv'

    def writerow(self, row, leaf=False):
        self.csv_writer.writerow([
            rec.encode('utf-8') if six.PY2 else rec
            for rec in row])

    def create_writer(self, headings, title=None):
        if six.PY2:
            self.content = io.BytesIO()
        else:
            self.content = io.StringIO()
        self.csv_writer = csv.writer(self.content)

    def flush_writer(self):
        self.content.seek(0)
        return self.content

    def get_filename(self):
        return datetime.datetime.now().strftime(self.basename + '-%Y%m%d.csv')


class SelfAssessmentXLSXView(SelfAssessmentSpreadsheetView):

    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def writerow(self, row, leaf=False):
        self.wsheet.append(row)
        if not leaf:
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.wsheet._current_row,
                    max_row=self.wsheet._current_row):
                row_cells[0].font = self.headingFont

    def create_writer(self, headings, title=None):
        SCALE = 11.9742857142857
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            self.wsheet.title = title
        self.wsheet.row_dimensions[1].height = 0.36 * (6 * SCALE)
        self.wsheet.column_dimensions['A'].width = 6.56 * SCALE
        for col_num in range(0, len(headings)):
            self.wsheet.column_dimensions[chr(ord('B') + col_num)].width \
                = 0.99 * SCALE
        self.headingFont = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF0071BB')

    def flush_writer(self):
        border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000'))
        alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=False,
            shrink_to_fit=False, indent=0)
        titleFill = PatternFill(fill_type=FILL_SOLID, fgColor='FFDDD9C5')
        subtitleFill = PatternFill(fill_type=FILL_SOLID, fgColor='FFEEECE2')
        subtitleFont = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        row = self.wsheet.row_dimensions[1]
        row.fill = titleFill
        row.font = Font(
            name='Calibri', size=20, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        row.alignment = alignment
        row.border = border
        self.wsheet.merge_cells('A1:F1')
        row = self.wsheet.row_dimensions[2]
        row.fill = subtitleFill
        row.font = subtitleFont
        row.alignment = alignment
        row.border = border
        row = self.wsheet.row_dimensions[3]
        row.fill = subtitleFill
        row.font = subtitleFont
        row.alignment = alignment
        row.border = border
        self.wsheet.merge_cells('B2:F2')
        self.wsheet.merge_cells('A2:A3')
        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def get_filename(self):
        return datetime.datetime.now().strftime(self.basename + '-%Y%m%d.xlsx')
