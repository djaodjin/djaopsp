# Copyright (c) 2020, DjaoDjin inc.
# see LICENSE.
from __future__ import unicode_literals

import csv, io, json, logging, re
from collections import namedtuple

from django.db import connection
from django.db.models import F, Count
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic.base import TemplateView
from deployutils.crypt import JSONEncoder
from deployutils.helpers import datetime_or_now, update_context_urls
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID
from pages.models import PageElement
from survey.models import Answer, Choice, Metric, Sample, Unit

from ..api.benchmark import PracticeBenchmarkMixin
from ..compat import reverse, six
from ..helpers import as_measured_value, get_testing_accounts
from ..mixins import ReportMixin, BestPracticeMixin
from ..models import ColumnHeader, Consumption, get_scored_answers
from ..scores import populate_account
from ..serializers import ConsumptionSerializer
from ..suppliers import get_supplier_managers
from ..templatetags.navactive import assessment_choices


LOGGER = logging.getLogger(__name__)


class AssessmentAnswer(object):

    def __init__(self, **kwargs):
        for key, val in six.iteritems(kwargs):
            setattr(self, key, val)

    def __getattr__(self, name):
        return getattr(self.consumption, name)


class AssessmentBaseMixin(PracticeBenchmarkMixin, BestPracticeMixin):
    # Implementation Note: uses BestPracticeMixin in order to display
    # bestpractice info through links in assess and improve pages.

    def decorate_visible_column_headers(self, root):
        # attach visible column headers
        hidden_columns = {}
        for icon_path, icon_tuple in six.iteritems(root[1]):
            hidden_columns[icon_path] = {}
            hidden = set([row['slug']
                for row in ColumnHeader.objects.filter(
                hidden=True, path=icon_path).values('slug')])
            value_headers = []
            for col_header in [
                    {"slug": "environmental_value",
                     "title": "/static/img/green-leaf.png",
                     "tooltip": "Environmental value"},
                    {"slug": "business_value",
                     "title": "/static/img/cogs.png",
                     "tooltip": "Ops/Maintenance value"},
                    {"slug": "profitability",
                     "title": "/static/img/dollar-sign.png",
                     "tooltip": "Financial value"},
                    {"slug": "implementation_ease",
                     "title": "/static/img/shovel.png",
                     "tooltip": "Implementation ease"},
                    {"slug": "avg_value",
                     "title": "Average Value"}]:
                if not col_header['slug'] in hidden:
                    value_headers += [col_header]
            icon_tuple[0]['value_headers'] = value_headers
            icon_tuple[0]['value_headers_len'] = len(value_headers) + 2
            icon_tuple[0]['colspan'] = icon_tuple[0]['value_headers_len']
        self._report_queries("attached visiblity of columns.")


class AssessmentView(AssessmentBaseMixin, TemplateView):

    template_name = 'envconnect/assess.html'
    breadcrumb_url = 'assess'

    def get_breadcrumb_url(self, path):
        organization = self.kwargs.get('organization', None)
        if organization:
            return reverse('assess_organization_redirect',
                args=(organization, path))
        return super(AssessmentView, self).get_breadcrumb_url(path)

    def get_context_data(self, **kwargs):
        context = super(AssessmentView, self).get_context_data(**kwargs)
        self.get_or_create_assessment_sample()
        segment_url, segment_prefix, segment_element = self.segment
        root = self.get_report_tree(load_text=(self.survey.slug == 'framework'))
        if root:
            self.decorate_visible_column_headers(root)
            context.update({
                'root': root,
                'entries': json.dumps(root, cls=JSONEncoder)
            })

        prev_samples = [(reverse('assess_organization',
            args=(self.account, prev_sample, self.kwargs.get('path'))),
                prev_sample.created_at)
            for prev_sample in Sample.objects.filter(
                is_frozen=True,
                extra__isnull=True,
                survey=self.survey,
                account=self.account).order_by('-created_at')]
        if prev_samples:
            context.update({'prev_samples': prev_samples})
            if self.sample.is_frozen:
                selected_sample = reverse('assess_organization',
                    args=(self.account, self.sample, self.kwargs.get('path')))
                context.update({'selected_sample': selected_sample})

        nb_questions = Consumption.objects.filter(
            path__startswith=segment_prefix).count()
        nb_required_questions = Consumption.objects.filter(
            path__startswith=segment_prefix,
            default_metric_id=self.default_metric_id).count()
        nb_answers = Answer.objects.filter(sample=self.sample,
            question__default_metric=F('metric_id'),
#            question__default_metric_id=self.default_metric_id,
            question__path__startswith=segment_prefix).count()
        nb_required_answers = Answer.objects.filter(sample=self.sample,
            question__default_metric=F('metric_id'),
            question__default_metric_id=self.default_metric_id,
            question__path__startswith=segment_prefix).count()
        context.update({
            'nb_answers': nb_answers,
            'nb_required_answers': nb_required_answers,
            'nb_questions': nb_questions,
            'nb_required_questions': nb_required_questions,
            'page_icon': self.icon,
            'sample': self.sample,
            'survey': self.sample.survey,
            'role': "tab",
            'score_toggle': self.request.GET.get('toggle', False)})

        # Find supplier managers subscribed to this profile
        # to share scorecard with.
        if self.manages(self.account):
            context.update({
                'supplier_managers': json.dumps(
                    get_supplier_managers(self.account))})

        organization = context['organization']
        update_context_urls(context, {
            'download': reverse(
                'assess_organization_sample_download',
                args=(organization, self.sample, segment_prefix)),
            'api_assessment_sample': reverse(
                'survey_api_sample', args=(organization, self.sample)),
            'api_assessment_freeze': reverse(
                'survey_api_sample_freeze', args=(organization, self.sample,
                segment_prefix)),
            'api_assessment_sample_new': reverse(
                'survey_api_sample_new', args=(organization,)),
            'api_scorecard_share': reverse('api_scorecard_share',
                args=(organization, segment_prefix)),
        })
        return context

    def get(self, request, *args, **kwargs):
        from_root, _ = self.breadcrumbs
        if not from_root or from_root == "/":
            return HttpResponseRedirect(reverse('homepage'))
        return super(AssessmentView, self).get(request, *args, **kwargs)


class AssessmentSpreadsheetView(AssessmentBaseMixin, TemplateView):

    basename = 'assessment'
    indent_step = '    '

    @staticmethod
    def _get_consumption(element):
        return element.get('consumption', None)

    @staticmethod
    def _get_tag(element):
        return element.get('tag', "")

    @staticmethod
    def _get_title(element):
        return element.get('title', "")

    def insert_path(self, tree, parts=None):
        if not parts:
            return tree
        if not parts[0] in tree:
            tree[parts[0]] = {}
        return self.insert_path(tree[parts[0]], parts[1:])

    def writerow(self, row, leaf=False):
        pass

    def write_tree(self, root, indent=''):
        """
        The *root* parameter looks like:
        (PageElement, [(PageElement, [...]), (PageElement, [...]), ...])
        """
        if not root[1]:
            # We reached a leaf
            consumption = self._get_consumption(root[0])
            row = [indent + self._get_title(root[0])]
            if consumption:
                for heading in self.get_headings(self._get_tag(root[0])):
                    if consumption.get('implemented', "") == heading:
                        row += ['X']
                    else:
                        row += ['']
                measures = consumption.get('measures', None)
                if measures:
                    comments = ""
                    sep = ""
                    for measure in measures:
                        if 'text' in measure and measure['text']:
                            comments += sep + str(measure['text'])
                        else:
                            comments += sep + str(measure['measured'])
                        sep = " "
                    if comments:
                        row += [comments]
            self.writerow(row, leaf=True)
        else:
            self.writerow([indent + self._get_title(root[0])])
            for element in six.itervalues(root[1]):
                self.write_tree(element, indent=indent + self.indent_step)

    def get(self, *args, **kwargs):
        #pylint: disable=unused-argument,too-many-locals
        # All assessment questions for a segment, regardless
        # of the actual from_path.
        segment_url, segment_prefix, segment_element = self.segment

        # `segment_prefix` is fully qualified path to segment.
        # We use cut=None here so we print out the full assessment
# XXX   root = self._build_tree(segment_element, segment_prefix, cut=None)
        root = self._build_tree(None, "", cut=None)

        self.headings = self.get_headings(self._get_tag(root[0]))
        self.create_writer(self.headings, title=self._get_title(root[0]))
        self.writerow(
            ["Assessment - Environmental practices"], leaf=True)
        self.writerow(
            ["Practice", "Implemented as a standard practice?"], leaf=True)
        self.writerow([""] + self.headings, leaf=True)
        indent = self.indent_step
        for nodes in six.itervalues(root[1]):
            self.writerow([indent + self._get_title(nodes[0])])
            for elements in six.itervalues(nodes[1]):
                self.write_tree(elements, indent=indent + self.indent_step)
        # Environmental metrics measured/reported
        measured_metrics = None
        # removed duplicate dump in .xslx spreadsheet when
        # `self.get_measured_metrics_context()` is used instead of `None`.
        if measured_metrics:
            self.writerow([])
            self.measured_title_row_idx = self.writerow(
                ["Environmental metrics measured/reported"], leaf=True)
            self.writerow(["Practice", "metric"]
                + ([""] * (len(self.headings) // 2))
                + ["measured"], leaf=True)
            for measured_metric in measured_metrics:
                look = re.match(r'.*indent-header-(\d+)', measured_metric[0])
                indent = " " * int(look.group(1))
                datapoints = measured_metric[3].get('environmental_metrics', [])
                self.writerow(["%s%s" % (indent, measured_metric[2].title)],
                    leaf=bool(datapoints))
                for datapoint in datapoints:
                    self.writerow(["",
                        datapoint['metric_title']]
                        + ([""] * (len(self.headings) // 2))
                        + [str(datapoint['measured'])], leaf=True)

        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = 'attachment; filename="{}"'.format(
            self.get_filename())
        return resp

    @staticmethod
    def get_headings(tag):
        return [str(choice) for choice in assessment_choices(tag)]


class AssessmentCSVView(AssessmentSpreadsheetView):

    content_type = 'text/csv'

    def writerow(self, row, leaf=False):
        self.csv_writer.writerow([
            rec.encode('utf-8') if six.PY2 else rec
            for rec in row])

    def create_writer(self, headings, title=None):
        #pylint:disable=unused-argument
        if six.PY2:
            self.content = io.BytesIO()
        else:
            self.content = io.StringIO()
        self.csv_writer = csv.writer(self.content)

    def flush_writer(self):
        self.content.seek(0)
        return self.content

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.csv')


class AssessmentXLSXView(AssessmentSpreadsheetView):

    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access
        self.wsheet.append(row)
        if leaf:
            if len(row) >= 6:
                for row_cells in self.wsheet.iter_rows(
                        min_row=self.wsheet._current_row,
                        max_row=self.wsheet._current_row):
                    row_cells[0].alignment = self.heading_alignment
                self.wsheet.row_dimensions[self.wsheet._current_row].height = 0
        else:
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.wsheet._current_row,
                    max_row=self.wsheet._current_row):
                row_cells[0].font = self.heading_font
                row_cells[0].alignment = self.heading_alignment
        return self.wsheet._current_row

    def create_writer(self, headings, title=None):
        col_scale = 11.9742857142857
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            # Prevents 'Invalid character / found in sheet title' errors
            self.wsheet.title = title.replace('/', '-')
        self.wsheet.row_dimensions[1].height = 0.36 * (6 * col_scale)
        self.wsheet.column_dimensions['A'].width = 6.56 * col_scale
        for col_num in range(0, len(headings)):
            self.wsheet.column_dimensions[chr(ord('B') + col_num)].width \
                = 0.99 * col_scale
        self.heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF0071BB')
        self.heading_alignment = Alignment(wrap_text=True)

    def flush_writer(self):
        border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000'))
        alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=False,
            shrink_to_fit=False, indent=0)
        title_fill = PatternFill(fill_type=FILL_SOLID, fgColor='FFDDD9C5')
        title_font = Font(
            name='Calibri', size=20, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        subtitle_fill = PatternFill(fill_type=FILL_SOLID, fgColor='FFEEECE2')
        subtitle_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        # Implementation Note: We style the cells here instead of the rows
        # otherwise opening the file on Excel leads to weird background coloring
        # (LibreOffice looks fine).
        cell = self.wsheet['A1']
        cell.fill = title_fill
        cell.font = title_font
        cell.border = border
        cell.alignment = alignment
        cell = self.wsheet['A2']
        cell.fill = subtitle_fill
        cell.font = subtitle_font
        cell.border = border
        cell.alignment = alignment
        cell = self.wsheet['B2']
        cell.fill = subtitle_fill
        cell.font = subtitle_font
        cell.border = border
        cell.alignment = alignment
        for col in ['B', 'C', 'D', 'E', 'F']:
            cell = self.wsheet['%s3' % col]
            cell.fill = subtitle_fill
            cell.font = subtitle_font
            cell.border = border
            cell.alignment = alignment
        self.wsheet.merge_cells('A1:F1')
        self.wsheet.merge_cells('B2:F2')
        self.wsheet.merge_cells('A2:A3')

        # Environmental metrics measured/reported
        if hasattr(self, 'measured_title_row_idx'):
            col_scale = 11.9742857142857
            self.wsheet.row_dimensions[self.measured_title_row_idx].height = (
                0.36 * (6 * col_scale))
            self.wsheet.merge_cells(
                start_row=self.measured_title_row_idx, start_column=1,
                end_row=self.measured_title_row_idx,
                end_column=1 + len(self.headings))
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.measured_title_row_idx,
                    max_row=self.measured_title_row_idx):
                row_cells[0].fill = title_fill
                row_cells[0].font = title_font
                row_cells[0].border = border
                row_cells[0].alignment = alignment

            for row_cells in self.wsheet.iter_rows(
                    min_row=self.measured_title_row_idx + 1,
                    max_row=self.measured_title_row_idx + 1):
                row_cells[0].fill = subtitle_fill
                row_cells[0].font = subtitle_font
                row_cells[0].alignment = alignment
                for cell in row_cells[1:]:
                    cell.fill = subtitle_fill
                    cell.font = subtitle_font
                    cell.border = border
                    cell.alignment = alignment
            self.wsheet.merge_cells(
                start_row=self.measured_title_row_idx + 1,
                start_column=2,
                end_row=self.measured_title_row_idx + 1,
                end_column=2 + len(self.headings) // 2)
            self.wsheet.merge_cells(
                start_row=self.measured_title_row_idx + 1,
                start_column=3 + len(self.headings) // 2,
                end_row=self.measured_title_row_idx + 1,
                end_column=3 + (len(self.headings) - len(self.headings) // 2))

        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')


class TargetsView(AssessmentView):
    """
    View that specifically filters the targets out of the assessment questions
    """
    breadcrumb_url = 'targets'


class CompleteView(AssessmentView, TemplateView):

    template_name = 'envconnect/complete.html'
    breadcrumb_url = 'complete'
