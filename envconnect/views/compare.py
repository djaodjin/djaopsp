# Copyright (c) 2020, DjaoDjin inc.
# see LICENSE.

import io, json, logging, re
from collections import OrderedDict

from dateutil.relativedelta import relativedelta
from deployutils.helpers import update_context_urls
from django.contrib.auth import get_user_model
from django.db import connection
from django.db.models import Count, Q
from django.utils import six
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView
from deployutils.apps.django.templatetags.deployutils_prefixtags import (
    site_prefixed)
from deployutils.helpers import datetime_or_now
from openpyxl import Workbook
from pages.models import PageElement
from survey.models import Matrix
from survey.views.matrix import MatrixDetailView

from ..compat import reverse
from ..api.benchmark import BenchmarkMixin
from ..api.dashboards import SupplierListMixin
from ..serializers import AccountSerializer
from ..helpers import as_valid_sheet_title
from ..mixins import AccountMixin, BreadcrumbMixin
from ..models import Consumption, _show_query_and_result

LOGGER = logging.getLogger(__name__)


class SuppliersView(AccountMixin, BreadcrumbMixin, TemplateView):

    template_name = 'envconnect/reporting/index.html'

    def get_context_data(self, **kwargs):
        context = super(SuppliersView, self).get_context_data(**kwargs)
        from_root, trail = self.breadcrumbs
        section_prefix = '/'.join([
            part for part in from_root.split('/') if part])
        update_context_urls(context, {
            'api_suppliers': reverse('api_suppliers',
                args=(self.account, section_prefix)),
            'api_accessibles': site_prefixed(
                "/api/profile/%(account)s/plans/%(account)s-report/"\
                "subscriptions/" % {'account': self.account}),
            'api_organizations': site_prefixed("/api/profile/"),
            'api_organization_profile': site_prefixed(
                "/api/profile/%(account)s/" % {'account': self.account}),
            'download': reverse('reporting_organization_download',
                                args=(self.account, from_root)),
            'improvements_download': reverse(
                'reporting_organization_improvements_download',
                args=(self.account, from_root))
        })
        try:
            extra = json.loads(self.account.extra)
        except (IndexError, TypeError, ValueError) as err:
            extra = {}
        start_at = extra.get('start_at', None)
        context.update({
            'score_toggle': True,
            'account_extra': self.account.extra,
            'date_range': {
                'start_at': start_at,
                'ends_at': (datetime_or_now() + relativedelta(days=1)
                ).isoformat(),
            }
        })
        return context


class PortfoliosDetailView(BenchmarkMixin, MatrixDetailView):

    matrix_url_kwarg = 'path'

    def get_available_matrices(self):
        return Matrix.objects.filter(account=self.account)

    def get_object(self, queryset=None):
        if queryset is None:
            queryset = self.get_queryset()
        candidate = self.kwargs.get(self.matrix_url_kwarg)
        if candidate.startswith('/'):
            candidate = candidate[1:]
        return get_object_or_404(queryset, slug=candidate)

    def get_context_data(self, **kwargs):
        #pylint:disable=too-many-locals,too-many-statements
        candidate = self.kwargs.get(self.matrix_url_kwarg)
        if candidate.startswith("/"):
            candidate = candidate[1:]
        parts = candidate.split("/")
        if len(parts) > 1:
            candidate = parts[0]

        context = MatrixDetailView.get_context_data(self, **kwargs)
        context.update({'organization': self.kwargs.get('organization')})
        context.update({'available_matrices': self.get_available_matrices()})

        try:
            PageElement.objects.get(slug=candidate)
        except PageElement.DoesNotExist:
            # It is not a breadcrumb path (ex: totals).
            #pylint:disable=unsubscriptable-object
            del self.kwargs[self.matrix_url_kwarg]

        from_root, trail = self.breadcrumbs
        segment_url, segment_prefix, segment_element = self.segment
        parts = from_root.split("/")
        if segment_element:
            root = self._build_tree(segment_element, from_root)
            self.decorate_with_breadcrumbs(root)
            # Remove sgement chart that would otherwise be added.
            charts = self.get_charts(root, excludes=[segment_element.slug])
        else:
            # totals
            charts = []
            segments = set([])
            for extra in self.account_queryset.filter(
                    subscription__plan__slug='%s-report' % self.account).values(
                        'extra'):
                try:
                    extra = extra.get('extra')
                    if extra:
                        extra = json.loads(extra.replace("'", '"'))
                        segments |= set([extra.get('industry')])
                except (IndexError, TypeError, ValueError) as _:
                    pass
            flt = None
            for segment in segments:
                if flt is None:
                    flt = Q(slug__startswith=segment)
                else:
                    flt |= Q(slug__startswith=segment)
            if True: #pylint:disable=using-constant-test
                # XXX `flt is None:` not matching the totals columns
                queryset = self.object.cohorts.all()
            else:
                queryset = self.object.cohorts.filter(flt)
            for cohort in queryset.order_by('title'):
                candidate = cohort.slug
                look = re.match(r"(\S+)(-\d+)$", candidate)
                if look:
                    candidate = look.group(1)
                element = PageElement.objects.filter(slug=candidate).first()
                charts += [{
                    'slug': cohort.slug,
                    'breadcrumbs': [cohort.title],
                    'icon': element.text if element is not None else "",
                    'icon_css': 'orange'
                }]
            charts = []
        url_kwargs = self.get_url_kwargs()
        url_kwargs.update({self.matrix_url_kwarg: self.object})
        for chart in charts:
            candidate = chart['slug']
            look = re.match(r"(\S+)(-\d+)$", candidate)
            if look:
                matrix_slug = '/'.join([look.group(1)])
            else:
                matrix_slug = '/'.join([str(self.object), candidate])
            url_kwargs.update({self.matrix_url_kwarg: matrix_slug})
            api_urls = {'matrix_api': reverse('matrix_api', kwargs=url_kwargs)}
            chart.update({'urls': api_urls})
        context.update({'charts': charts})

        update_context_urls(context, {
            'reporting_organization': reverse('reporting_organization',
                args=(self.account, ''))
        })
        return context


class SuppliersSummaryXLSXView(SupplierListMixin, TemplateView):
    """
    Download scores of all reporting entities as an Excel spreadsheet.
    """
    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    basename = 'dashboard'

    headings = [
        'Supplier name', 'Categories',
        'Contact name', 'Contact email', 'Contact phone',
        'Last activity', 'Status', 'Industry segment', 'Score',
        '# N/A', 'Reporting publicly', 'Environmental fines',
        '# Planned actions', 'Reported measurements', 'Targets',
#        'Full-time Employee Count', 'Annual Revenue (USD)'
#        'Responded in 2018', 'Responded in 2019'
    ]

    def get_headings(self):
        return self.headings

    @staticmethod
    def get_indent_bestpractice(depth=0):
        return "  " * depth

    @staticmethod
    def get_indent_heading(depth=0):
        return "  " * depth

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')

    def _writerecord(self, rec, categories, reporting_status, last_activity_at,
                     report_to=None):
        #pylint:disable=too-many-arguments,too-many-locals
        if rec.get('requested_at'):
            normalized_score = "Requested"
            segment = "Requested"
            nb_na_answers = "Requested"
            reporting_publicly = "Requested"
            reporting_fines = "Requested"
            nb_planned_improvements = "Requested"
            measurements = "Requested"
            targets = "Requested"
            employee_count = "Requested"
            revenue_generated = "Requested"
        else:
            normalized_score = rec.get('normalized_score', "N/A")
            segment = rec.get('segment', "N/A")
            nb_na_answers = rec.get('nb_na_answers', "N/A")
            reporting_publicly = rec.get('reporting_publicly', "N/A")
            if reporting_publicly and str(reporting_publicly) != "N/A":
                reporting_publicly = "Yes"
            reporting_fines = rec.get('reporting_fines', "N/A")
            if reporting_fines and str(reporting_fines) != "N/A":
                reporting_fines = "Yes"
            nb_planned_improvements = rec.get('nb_planned_improvements', "N/A")
            measurements = '\n'.join(rec.get('reported', []))
            targets = '\n'.join(rec.get('targets', []))
            employee_count = rec.get('employee_count', "N/A")
            revenue_generated = rec.get('revenue_generated', "N/A")
        contact_name = rec.get('contact_name', "")
        contact_phone = rec.get('phone', "")
        if not contact_name:
            contact_model = get_user_model()
            try:
                contact = contact_model.objects.get(
                    email__iexact=rec.get('email', ""))
                contact_name = contact.get_full_name()
            except contact_model.DoesNotExist as err:
                LOGGER.warning("supplier '%s', contact e-mail '%s' not found!",
                    rec['printable_name'], rec.get('email', ""))
        self.wsheet.append([
            rec['printable_name'], categories,
            contact_name,
            rec.get('email', ""),
            contact_phone,
            last_activity_at, reporting_status,
            segment, normalized_score,
            nb_na_answers,
            reporting_publicly, reporting_fines,
            nb_planned_improvements, measurements, targets])
            #XXX employee_count, revenue_generated,
            #XXX report_to if report_to else ""])
#            "Yes" if rec.get('improvement_completed') else "",
#            "Yes" if rec.get('assessment_completed') else ""])

    def writerow(self, rec, headings=None):
        last_activity_at = rec.get('last_activity_at', "")
        if last_activity_at:
            last_activity_at = last_activity_at.strftime("%Y-%m-%d")
        reporting_status = rec.get('reporting_status')
        if reporting_status < len(AccountSerializer.REPORTING_STATUS):
            reporting_status = (
                AccountSerializer.REPORTING_STATUS[reporting_status][1])
        else:
            reporting_status = ""
        extra = rec['extra']
        if extra and isinstance(extra, six.string_types):
            try:
                extra = json.loads(extra)
            except (TypeError, ValueError):
                extra = {}
        categories = ','.join(extra.keys()) if extra else ""
        if headings:
            self._writerecord(rec, categories, reporting_status,
                last_activity_at)
        else:
            for rep in [(self.account.slug, self.account.full_name)]:
                report_to = "" if rep[0] == self.account.slug else rep[1]
                self._writerecord(rec, categories, reporting_status,
                    last_activity_at, report_to=report_to)
                segment_slug = None
                if segment_slug:
                    if segment_slug not in self.suppliers_per_segment:
                        self.suppliers_per_segment[segment_slug] = set(
                            [])
                    self.suppliers_per_segment[segment_slug] |= set(
                        [rec['slug']])

    def get(self, request, *args, **kwargs):
        #pylint: disable=unused-argument,too-many-locals,too-many-nested-blocks
        #pylint: disable=too-many-statements
        rollup_tree = self.rollup_scores(force_score=True)
        self.suppliers_per_segment = {}
        wbook = Workbook()

        # Populate the Total sheet
        self.wsheet = wbook.active
        self.wsheet.title = as_valid_sheet_title("Suppliers invited to TSP")
        self.wsheet.append(["Utility member", self.account.full_name])
        contact_model = get_user_model()
        contact_name = ""
        try:
            contact = contact_model.objects.get(
                email__iexact=self.account.email)
            contact_name = contact.get_full_name()
        except contact_model.DoesNotExist as err:
            LOGGER.warning("member '%s', contact e-mail '%s' not found!",
                self.account.full_name, self.account.email)
        self.wsheet.append(["Utility contact name", contact_name])
        self.wsheet.append(["Utility contact email", self.account.email])
        self.wsheet.append(["Utility contact phone", self.account.phone])
        self.wsheet.append([])
        self.wsheet.append(self.get_headings())
        for account in self.get_suppliers(rollup_tree):
            self.writerow(account)

        # Populate per-segment sheets
        if False:
            for container in six.itervalues(rollup_tree[1]):
                segment_suppliers = self.suppliers_per_segment.get(
                    container[0]['slug'], set([]))
                if not segment_suppliers:
                    continue
                for segment in six.itervalues(container[1]):
                    headings = [val[0]['title']
                        for val in six.itervalues(segment[1])]
                    all_headings = self.get_headings()[:-1] + headings
                    suppliers_per_segment = self.get_suppliers(segment)
                    if suppliers_per_segment:
                        self.wsheet = wbook.create_sheet(
                            title=as_valid_sheet_title(segment[0]['title']))
                        self.wsheet.append(all_headings)
                        for account in suppliers_per_segment:
                            if account['slug'] in segment_suppliers:
                                self.writerow(account, headings=headings)

        # Prepares the result file
        content = io.BytesIO()
        wbook.save(content)
        content.seek(0)

        resp = HttpResponse(content, content_type=self.content_type)
        resp['Content-Disposition'] = 'attachment; filename="{}"'.format(
            self.get_filename())
        return resp


class SuppliersAssessmentsXLSXView(SupplierListMixin, TemplateView):
    """
    Download detailed answers of each suppliers
    """
    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    basename = 'dashboard-assessments'

    indent_step = '    '

    @staticmethod
    def _get_consumption(element):
        return element.get('consumption', None)

    @staticmethod
    def _get_tag(element):
        return element.get('tag', "")

    @staticmethod
    def _get_title(element):
        return element.get('title', "")

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access
        self.wsheet.append(row)

    def write_tree(self, root, indent=''):
        """
        The *root* parameter looks like:
        (PageElement, [(PageElement, [...]), (PageElement, [...]), ...])
        """
        if not root[1]:
            # We reached a leaf
            row = [indent + self._get_title(root[0])]
            by_accounts = self.by_paths.get(root[0]['path'])
            if by_accounts:
                for account_id in self.requested_accounts_pk:
                    answer = by_accounts[account_id]
                    text = answer['measured']
                    if answer['comments']:
                        text += " - %s" % answer['comments']
                    row += [text]
            self.writerow(row, leaf=True)
        else:
            self.writerow([indent + self._get_title(root[0])])
            for element in six.itervalues(root[1]):
                self.write_tree(element, indent=indent + self.indent_step)

    def get_latest_samples(self, from_root):
        return Consumption.objects.get_latest_samples_by_prefix(
            before=self.ends_at, prefix=from_root)

    def get(self, request, *args, **kwargs):
        #pylint:disable=too-many-statements,too-many-locals
        from_root, trail = self.breadcrumbs
        head = trail[-1][0]
        # We use cut=None here so we print out the full assessment
        root = self._build_tree(head, from_root, cut=None)

        latest_assessments = self.get_latest_samples(from_root)

        reporting_answers_sql = """
WITH samples AS (
    %(latest_assessments)s
)
SELECT
    survey_question.path,
    samples.account_id,
    survey_choice.text,
    survey_answer.metric_id,
    survey_answer.unit_id
FROM survey_answer
INNER JOIN survey_metric
  ON survey_answer.metric_id = survey_metric.id
INNER JOIN survey_choice
  ON survey_choice.unit_id = survey_metric.unit_id
INNER JOIN survey_question
  ON survey_answer.question_id = survey_question.id
INNER JOIN samples
  ON survey_answer.sample_id = samples.id
WHERE survey_choice.id = survey_answer.measured AND
  samples.account_id IN %(accounts)s""" % {
            'latest_assessments': latest_assessments,
            'accounts': self.requested_accounts_pk_as_sql
        }
        _show_query_and_result(reporting_answers_sql)
        self.by_paths = {}
        with connection.cursor() as cursor:
            cursor.execute(reporting_answers_sql, params=None)
            for row in cursor:
                path = row[0]
                account_id = row[1]
                measured = row[2]
                metric_id = row[3]
                unit_id = row[4]
                if path not in self.by_paths:
                    by_accounts = OrderedDict()
                    for account in self.requested_accounts_pk:
                        by_accounts[account] = {'measured': "", 'comments': ""}
                    self.by_paths[path] = by_accounts
                if metric_id == self.default_metric_id:
                    self.by_paths[path][account_id]['measured'] = measured
                else:
                    self.by_paths[path][account_id]['comments'] += measured

        # Populate the worksheet
        wbook = Workbook()
        self.wsheet = wbook.active
        self.wsheet.title = as_valid_sheet_title("Answers")
        headings = ['']
        for account in self.requested_accounts:
            headings += [account.printable_name]
        self.wsheet.append(headings)

        by_accounts = OrderedDict()
        for account in self.requested_accounts_pk:
            by_accounts[account] = ""
        with connection.cursor() as cursor:
            cursor.execute(latest_assessments, params=None)
            for sample in cursor:
                last_activity_at = sample[2]
                account_id = sample[4]
                by_accounts[account_id] = last_activity_at
        headings = ['Last completed at']
        for account_id in self.requested_accounts_pk:
            last_activity_at = by_accounts[account_id]
            if last_activity_at:
                headings += [last_activity_at.strftime("%Y-%m-%d")]
            else:
                headings += [""]
        self.wsheet.append(headings)

        indent = self.indent_step
        for nodes in six.itervalues(root[1]):
            self.writerow([indent + self._get_title(nodes[0])])
            for elements in six.itervalues(nodes[1]):
                self.write_tree(elements, indent=indent + self.indent_step)

        # Prepares the result file
        content = io.BytesIO()
        wbook.save(content)
        content.seek(0)

        resp = HttpResponse(content, content_type=self.content_type)
        resp['Content-Disposition'] = 'attachment; filename="{}"'.format(
            self.get_filename())
        return resp

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')


class SuppliersPlanningXLSXView(SuppliersAssessmentsXLSXView):
    """
    Download detailed planned improvements of each suppliers
    """
    basename = 'dashboard-planning'

    def get_latest_samples(self, from_root):
        return Consumption.objects.get_latest_samples_by_prefix(
            before=self.ends_at, prefix=from_root, tag='is_planned')


class SuppliersImprovementsXLSXView(SupplierListMixin, TemplateView):
    """
    Download planned actions accross reporting entities as an Excel spreadsheet.
    """
    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    basename = 'dashboard-improvements'

    @staticmethod
    def get_indent_bestpractice(depth=0):
        return "  " * depth

    @staticmethod
    def get_indent_heading(depth=0):
        return "  " * depth

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')

    def get(self, request, *args, **kwargs):
        #pylint: disable=unused-argument,too-many-locals
        wbook = Workbook()

        # Populate improvements planned sheet
        practices = Consumption.objects.filter(
            answer__sample__extra='is_planned',
            answer__sample__is_frozen=True,
            answer__metric_id=self.default_metric_id,
            answer__measured=Consumption.NEEDS_SIGNIFICANT_IMPROVEMENT
        ).annotate(nb_suppliers=Count('answer__sample')).values(
            'path', 'nb_suppliers')

        depth = 2
        root = (OrderedDict({}), OrderedDict({}))
        for practice in practices:
            node = self._insert_path(root, practice['path'], depth=depth)
            node[0].update({'nb_suppliers': practice['nb_suppliers']})
        root = self._natural_order(root)
        rows = self.flatten_answers(root, '')

        self.wsheet = wbook.active
        self.wsheet.title = as_valid_sheet_title("Improvements")
        self.wsheet.append(('Best practice',
            'Nb suppliers who have selected the practice for improvement.'))
        for row in rows:
            indent = row[0]
            title = row[2].title
            nb_suppliers = row[3].get('nb_suppliers', "")
            self.wsheet.append([indent + title, nb_suppliers])

        # Prepares the result file
        content = io.BytesIO()
        wbook.save(content)
        content.seek(0)

        resp = HttpResponse(content, content_type=self.content_type)
        resp['Content-Disposition'] = 'attachment; filename="{}"'.format(
            self.get_filename())
        return resp
