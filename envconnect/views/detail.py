# Copyright (c) 2019, DjaoDjin inc.
# see LICENSE.

import io, json

from deployutils.helpers import datetime_or_now
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic import TemplateView, ListView
from django.utils import six
from lxml import html
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.fills import FILL_SOLID

from ..compat import reverse
from ..helpers import as_valid_sheet_title
from ..mixins import BreadcrumbMixin, BestPracticeMixin
from ..models import ColumnHeader


class DetailView(BestPracticeMixin, TemplateView):

    template_name = 'envconnect/detail.html'
    breadcrumb_url = 'summary'

    def get(self, request, *args, **kwargs):
        self._start_time()
        _, trail = self.breadcrumbs
        if not trail:
            return HttpResponseRedirect(reverse('homepage'))
        result = super(DetailView, self).get(request, *args, **kwargs)
        self._report_queries("request ready to send.")
        return result

    def get_breadcrumb_url(self, path):
        organization = self.kwargs.get('organization', None)
        if organization:
            return reverse('summary_organization_redirect', args=(organization, path))
        return super(DetailView, self).get_breadcrumb_url(path)

    def get_context_data(self, **kwargs):
        #pylint:disable=too-many-locals
        context = super(DetailView, self).get_context_data(**kwargs)
        from_root, trail = self.breadcrumbs
        # It is OK here to index by -1 since we would have generated a redirect
        # in the `get` method when the path is "/".
        root = self._build_tree(trail[-1][0], from_root)
        self._report_queries("content tree built.")

        # Removes framework additional questions and other assessments
        # questions which are irrelevant when making an improvement plan.
        to_remove = []
        for key, node in six.iteritems(root[1]):
            tags = node[0].get('tag', "")
            if tags and 'metrics' in tags:
                to_remove += [key]
        for key in to_remove:
            del root[1][key]

        # attach visible column headers
        hidden_columns = {}
        is_content_manager = context.get('is_content_manager', False)
        for icon_path, icon_tuple in six.iteritems(root[1]):
            hidden_columns[icon_path] = {}
            hidden = set([row['slug']
                for row in ColumnHeader.objects.filter(
                hidden=True, path=icon_path).values('slug')])
            profitability_headers = []
            for col_header in [
                        {"slug": "avg_energy_saving",
                         "title": "Savings",
                         "tooltip": "The average estimated percentage saved"\
" relative to spend in the area of focus (e.g. site energy, waste) resulting"\
" from the implementation of a best practice.\n***** > 5%\n****   3-5%\n***"\
"     2-3%\n**       1-2%\n*         < 1%\n Click individual best practice"\
" headings and navigate to the \"References\" section for more detail on data"\
" provenance."},
                        {"slug": "capital_cost",
                         "title": "Cost",
                        "tooltip": "The average estimated percentage of cost"\
" relative to spend in the area of focus (e.g. site energy, waste) to"\
" implement a best practice.\n$$$$$ < 10%\n$$$$   5-10%\n$$$     2-5%\n$$"\
"       1-2%\n$         < 1%\nClick individual best practice headings and"\
" navigate to the \"References\" section for more detail on data provenance."},
                        {"slug": "payback_period",
                         "title": "Payback years",
                         "tooltip": "Range: The range of payback values are"\
" calculated for a best practice implemented at a facility by the following"\
" formula: Simple Year-of-Payback = (Implementation Cost) / (Total Energy"\
" Cost Savings)\n\nAverage (in parentheses): The average represents an"\
" average of the totals represented by the range.\n\nClick individual best"\
" practice headings and navigate to the \"References\" section for more"\
" detail on data provenance."}]:
                if is_content_manager:
                    profitability_headers += [col_header]
                    hidden_columns[icon_path][col_header['slug']] = (
                        col_header['slug'] in hidden)
                elif not col_header['slug'] in hidden:
                    profitability_headers += [col_header]
            icon_tuple[0]['profitability_headers'] = profitability_headers
            icon_tuple[0]['profitability_headers_len'] = len(
                profitability_headers) + 1
            value_headers = []
            for col_header in [
                    {"slug": "environmental_value",
                     "title": "/static/img/green-leaf.png",
                     "tooltip": "Environmental value"},
                    {"slug": "business_value",
                     "title": "/static/img/cogs.png",
                     "tooltip": "Ops/Maintenance value"},
                    {"slug": "profitability",
                     "title": "/static/img/dollar-sign.png",
                     "tooltip": "Financial value"},
                    {"slug": "implementation_ease",
                     "title": "/static/img/shovel.png",
                     "tooltip": "Implementation ease"},
                    {"slug": "avg_value",
                     "title": "Average Value"}]:
                if is_content_manager:
                    value_headers += [col_header]
                    hidden_columns[icon_path][col_header['slug']] = (
                        col_header['slug'] in hidden)
                elif not col_header['slug'] in hidden:
                    value_headers += [col_header]
            icon_tuple[0]['value_headers'] = value_headers
            icon_tuple[0]['value_headers_len'] = len(value_headers) + 2
            icon_tuple[0]['colspan'] = max(
                icon_tuple[0]['profitability_headers_len'],
                icon_tuple[0]['value_headers_len'])
        self._report_queries("attached visiblity of columns.")

        if not is_content_manager:
            context.update({'sort_by': "{'agv_value': 'desc'}"})
        context.update({
            'role': "tab",
            'root': root,
            'entries': json.dumps(root),
            'hidden': json.dumps(hidden_columns)
        })
        return context


class DetailSpreadsheetView(BreadcrumbMixin, ListView):

    basename = 'best-practices'
    headings = ['Environmental', 'Ops/maintenance', 'Financial',
        'Implementation ease', 'AVERAGE VALUE']

    def create_writer(self, headings, title=None):
        raise NotImplemented(
            "`create_writer` must be implemented in derived class")

    def flush_writer(self):
        raise NotImplemented(
            "`flush_writer` must be implemented in derived class")

    def tree_depth(self, root):
        if not root:
            return 0
        element = root[0].get('consumption', None)
        if element:
            return 1
        depth = 0
        element = root[0]
        for _, nodes in six.iteritems(root[1]):
            depth = max(self.tree_depth(nodes), depth)
        return depth + 1

    def write_tree(self, root, indent=[]):
        if not root:
            return
        element = root[0].get('consumption', None)
        local_indent = indent
        if element:
            # We reached a leaf
            if len(local_indent) < self.depth:
                local_indent = local_indent + ["" for unnamed in range(
                    len(local_indent), self.depth - 1)] + [root[0]['title']]
            self.writerow(local_indent + [
                element['environmental_value'],
                element['business_value'],
                element['profitability'],
                element['implementation_ease'],
                element['avg_value']
            ], leaf=True)
        else:
            for _, nodes in six.iteritems(root[1]):
                self.write_tree(nodes, indent=local_indent + [root[0]['title']])

    def get(self, request, *args, **kwargs): #pylint: disable=unused-argument
        from_root, trail = self.breadcrumbs
        # It is OK here to index by -1 since we would have generated a redirect
        # in the `get` method when the path is "/".
        root = self._build_tree(trail[-1][0], from_root)
        self.depth = self.tree_depth(root)
        self.create_writer(self.get_headings(), title="Best practices")
        self.writerow(["Find value assignment at"\
" https://tspproject.org/docs/faq/#scorecard-1"], leaf=True)
        self.writerow(self.get_headings(), leaf=True)
        self.write_tree(root)
        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = \
            'attachment; filename="{}"'.format(self.get_filename())
        return resp

    def get_headings(self):
        return ['' for unnamed in range(0, self.depth)] + self.headings

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.csv')


class DetailXLSXView(DetailSpreadsheetView):
    """
    Downloads the best practices as an Excel spreadsheet.
    """

    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    valueFills = [
        PatternFill(fill_type=FILL_SOLID, fgColor='FF9CD76B'), # green-level-0
        PatternFill(fill_type=FILL_SOLID, fgColor='FF69B02B'), # green-level-1
        PatternFill(fill_type=FILL_SOLID, fgColor='FF007C3F'), # green-level-2
        PatternFill(fill_type=FILL_SOLID, fgColor='FFFFD700'), # green-level-3
    ]

    def create_writer(self, headings, title=None):
        col_scale = 11.9742857142857
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            self.wsheet.title = as_valid_sheet_title(title)
        self.wsheet.row_dimensions[2].height = 35
        for col_num in range(self.depth, self.depth + len(self.headings)):
            col_idx = chr(ord('A') + col_num)
            self.wsheet.column_dimensions[col_idx].width = 1.1 * col_scale
        self.heading_font = Font(
            name='Calibri', size=12, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF0071BB')
        self.heading_alignment = Alignment(wrap_text=True)
        self.border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000'))
        self.text_center = Alignment(horizontal='center')

    def flush_writer(self):
        #pylint:disable=protected-access,too-many-statements
        col_scale = 0.8
        bold_font = Font(
            name='Calibri', size=11, bold=True, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=False,
            shrink_to_fit=False, indent=0)
        subtitle_fill = PatternFill(fill_type=FILL_SOLID, fgColor='FFEEECE2')
        subtitle_font = Font(
            name='Calibri', size=10, bold=False, italic=False,
            vertAlign='baseline', underline='none', strike=False,
            color='FF000000')
        subtitle_alignment = Alignment(
            horizontal='center', vertical='center',
            text_rotation=0, wrap_text=True,
            shrink_to_fit=True, indent=0)
        for col_num in range(1, self.depth + 1):
            max_length = 0
            for row in self.wsheet.iter_rows(
                    min_row=3, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(len(cell.value), max_length)
            if max_length > 60:
                max_length = 60
            col_idx = chr(ord('A') + col_num - 1)
            self.wsheet.column_dimensions[col_idx].width = \
                max_length * col_scale

        # Implementation Note: We style the cells here instead of the rows
        # otherwise opening the file on Excel leads to weird background coloring
        # (LibreOffice looks fine).
        for col in [chr(ord('A') + x) for x in range(0,
                        self.depth + len(self.headings))]:
            cell = self.wsheet['%s2' % col]
            cell.fill = subtitle_fill
            cell.font = subtitle_font
            cell.border = self.border
            cell.alignment = subtitle_alignment

        # Write out the Excel file.
        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')

    def get_value_fill(self, val):
        if val < len(self.valueFills):
            return self.valueFills[val]
        return self.valueFills[-1]

    def writerow(self, row, leaf=False):
        #pylint:disable=protected-access
        self.wsheet.append(row)
        if leaf:
            bp_length = self.depth + len(self.headings)
            if len(row) >= bp_length:
                for row_cells in self.wsheet.iter_rows(
                        min_row=self.wsheet._current_row,
                        max_row=self.wsheet._current_row):
                    for cell_idx in range(self.depth, bp_length):
                        # environmental_value, business_value,
                        # implementation_ease, profitability,
                        # avg_value
                        try:
                            row_cells[cell_idx].border = self.border
                            row_cells[cell_idx].fill = self.get_value_fill(
                                int(row_cells[cell_idx].value))
                        except ValueError:
                            # might be the header
                            pass
        else:
            for row_cells in self.wsheet.iter_rows(
                    min_row=self.wsheet._current_row,
                    max_row=self.wsheet._current_row):
                row_cells[0].font = self.heading_font
                row_cells[0].alignment = self.heading_alignment


class ContentDetailXLSXView(DetailSpreadsheetView):
    """
    Outputs text content of best practices in an Excel spreadsheet.
    """
    basename = 'best-practices-content'
    headings = ['Description content']
    content_type = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def create_writer(self, headings, title=None):
        self.wbook = Workbook()
        self.wsheet = self.wbook.active
        if title:
            self.wsheet.title = as_valid_sheet_title(title)

    def flush_writer(self):
        # Write out the Excel file.
        content = io.BytesIO()
        self.wbook.save(content)
        content.seek(0)
        return content

    def writerow(self, row, leaf=False):
        self.wsheet.append(row)

    def write_tree(self, root, indent=[]):
        if not root:
            return
        title = root[0]['title']
        text = root[0].get('text', "")
        if text:
            text = html.fromstring(text).text_content()
        element = root[0].get('consumption', None)
        local_indent = indent
        if element:
            # We reached a leaf
            if len(local_indent) < self.depth:
                local_indent = local_indent + ["" for unnamed in range(
                    len(local_indent), self.depth - 1)] + [title]
            self.writerow(local_indent + [
                text
            ], leaf=True)
        else:
            for _, nodes in six.iteritems(root[1]):
                self.write_tree(nodes, indent=local_indent + [title])

    def get(self, request, *args, **kwargs): #pylint: disable=unused-argument
        from_root, trail = self.breadcrumbs
        # It is OK here to index by -1 since we would have generated a redirect
        # in the `get` method when the path is "/".
        root = self._build_tree(trail[-1][0], from_root, load_text=True)
        self.depth = self.tree_depth(root)
        self.create_writer(self.get_headings(), title="Best practices")
        self.writerow(self.get_headings(), leaf=True)
        self.write_tree(root)
        resp = HttpResponse(self.flush_writer(), content_type=self.content_type)
        resp['Content-Disposition'] = \
            'attachment; filename="{}"'.format(self.get_filename())
        return resp

    def get_filename(self):
        return datetime_or_now().strftime(self.basename + '-%Y%m%d.xlsx')
